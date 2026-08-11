import dbf
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import StaleElementReferenceException
import time
import tempfile
import io
import sys
import json
import datetime
import argparse
import zipfile
import threading
import queue
import xml.etree.ElementTree as ET
import pandas as pd
import openpyxl
from PIL import Image, ImageOps
import ddddocr
import logging
import os
import requests
import shutil
import pythoncom
import win32com.server.register
import calendar
import re
import unicodedata
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.common.by import By

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')    ##errors='replace' đảm bảo không bao giờ vỡ — ký tự lạ thành ?.
# Cấu hình log
# Vá #7 (04/08): trước đây ghim r'C:\test\com_server_error.log'. Dòng này chạy ngay
# lúc import, nên máy nào không sẵn thư mục C:\test là script chết vì FileNotFoundError
# TRƯỚC cả khi đọc tham số. Nay đặt cạnh chính file script (đè được bằng biến môi
# trường KT2000_LOG_DIR) và tự tạo thư mục — chuyển máy nào cũng chạy, không phụ
# thuộc ổ đĩa nào tồn tại.
_LOG_DIR = os.environ.get("KT2000_LOG_DIR") or os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "logs")
try:
    os.makedirs(_LOG_DIR, exist_ok=True)
except Exception:
    _LOG_DIR = tempfile.gettempdir()

logging.basicConfig(
    filename=os.path.join(_LOG_DIR, 'com_server_error.log'),
    level=logging.ERROR,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# =========================
# RUN LOG + HEARTBEAT HELPERS
# =========================
def ensure_parent_dir(path):
    if path:
        os.makedirs(os.path.dirname(path), exist_ok=True)

def append_run_log(log_path, message):
    try:
        ensure_parent_dir(log_path)
        ts = datetime.datetime.now().isoformat(timespec="seconds")
        with open(log_path, "a", encoding="utf-8", errors="replace") as f:
            f.write(f"[{ts}] {message}\n")
            f.flush()
            os.fsync(f.fileno())
    except Exception:
        pass

def short_exc(e):
    try:
        return f"{type(e).__name__}: {e}"
    except Exception:
        return str(e)

# Lấy danh sách hóa đơn từ chính file Excel danh sách vừa tải, thay cho việc gọi
# API search phân trang. Bản 1.3 đã làm như vậy; bản 2.0 lỡ bỏ mất nên mỗi tháng
# lại bắn thêm hàng chục request search — vừa chậm vừa dễ ăn 429.
# Excel đã có đủ 4 trường cần để dựng URL tải XML, không phải hỏi lại cổng.
DUNG_EXCEL_THAY_SEARCH = True


def build_ds_hd_tu_excel(excel_path, header_row=6):
    """
    Doc 4 truong can de tai XML tu file Excel danh sach (export-excel cua GDT):
      Ky hieu mau so -> khmshdon, Ky hieu hoa don -> khhdon,
      So hoa don -> shdon, MST nguoi ban -> nbmst.
    Tra ve list dict giong cau truc 'datas' cua search (du cho tai_hd dung).
    Dedup theo (khhdon, shdon).
    """
    def _norm(x):
        return re.sub(r"\s+", " ", str(x)).strip().lower() if x is not None else ""

    want = {
        # "ký" CÓ dấu sắc — header that la "Ky hieu mau so" viet co dau day du. Ban port
        # dau tien lam roi dau o chu nay, the la khong khop cot nao, build_ds_hd_tu_excel
        # luon nem loi va luon quay ve search: ca hai cai loi (bo search, va tang dan)
        # deu am tham khong chay, ma nhin ben ngoai khong thay gi khac.
        "khmshdon": ["ky hieu mau so", "ký hiệu mẫu số"],
        "khhdon":   ["ky hieu hoa don", "ký hiệu hóa đơn"],
        "shdon":    ["so hoa don", "số hóa đơn"],
        "nbmst":    ["mst nguoi ban/mst nguoi xuat hang",
                     "mst người bán/mst người xuất hàng"],
    }
    want_norm = {k: [_norm(x) for x in v] for k, v in want.items()}

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    try:
        ws = wb.active
        idx = {}
        for c in range(1, ws.max_column + 1):
            h = _norm(ws.cell(header_row, c).value)
            for key, cands in want_norm.items():
                if key not in idx and h in cands:
                    idx[key] = c

        missing = [k for k in want if k not in idx]
        if missing:
            raise ValueError(f"Excel thieu cot cho: {missing}")

        def _txt(r, c):
            v = ws.cell(r, c).value
            if v is None:
                return ""
            t = str(v).strip()
            if re.fullmatch(r"\d+\.0+", t):   # Excel luu so -> bo duoi .0
                t = t.split(".", 1)[0]
            return t

        ds_hd, seen = [], set()
        for r in range(header_row + 1, ws.max_row + 1):
            khhdon = _txt(r, idx["khhdon"])
            shdon = _txt(r, idx["shdon"])
            if not khhdon or not shdon:
                continue
            key = (khhdon, shdon)
            if key in seen:
                continue
            seen.add(key)
            ds_hd.append({
                "khmshdon": _txt(r, idx["khmshdon"]),
                "khhdon": khhdon,
                "shdon": shdon,
                "nbmst": _txt(r, idx["nbmst"]),
            })
        return ds_hd
    finally:
        wb.close()


# "Khong ton tai ho so goc" (HTTP 500) la ca HOP LE: hoa don dien/vien thong/ngan
# hang chi co trong Excel, cong khong giu ban goc. Gop chung voi 429/504 thi lan
# nao cung thay "loi" va khong ai biet cai nao dang di tim.
def la_khong_co_goc(lydo):
    t = str(lydo)
    return ("hồ sơ gốc" in t) or ("ho so goc" in t.lower())


# =========================
# STATUS + EVENTS (NEW)
# =========================
class StatusWriter:
    def __init__(self, path):
        self.path = path

    def write(self, data):
        data["updated_at"] = datetime.datetime.now().isoformat(timespec="seconds")
        tmp = self.path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
            f.flush()
            os.fsync(f.fileno())
        for i in range(5):
            try:
                os.replace(tmp, self.path)
                return
            except PermissionError:
                time.sleep(0.5)
        os.replace(tmp, self.path)


class EventWriter:
    def __init__(self, path):
        self.path = path

    def log(self, event, **kwargs):
        data = {
            "ts": datetime.datetime.now().isoformat(timespec="seconds"),
            "event": event,
            **kwargs
        }
        with open(self.path, "a", encoding="utf-8") as f:
            f.write(json.dumps(data, ensure_ascii=False) + "\n")
def make_job_paths(base_dir, job_id, MA_DONVI):
    # job_dir = os.path.join(base_dir, MA_DONVI, job_id)
    # Va #7a: chen tang NAM<nam> (doc nam tu job_id dang T{m}_{yyyy}_...)
    # Chi lay TEN job, bo moi thanh phan thu muc nguoi goi lo kem vao. Backend .NET
    # truyen job_id = "NAM2026\\T5_2026_HUY_THANH" (no tu ghep tang NAM de dung
    # duong dan status/stage), ma ham nay CUNG tu chen "NAM<nam>" -> ra cay long
    # nhau NAM2026\\NAM2026\\T5_... Ket qua: run.log lac sang cay rong, con raw/
    # outputs nam o cay dung, nhin thu muc thay hai ban giong het nhau.
    job_id = os.path.basename(str(job_id).replace("\\", "/").rstrip("/"))

    m_nam = re.search(r'_(20\d{2})_', str(job_id) + "_")
    nam_dir = f"NAM{m_nam.group(1)}" if m_nam else ""
    job_dir = os.path.join(base_dir, MA_DONVI, nam_dir, job_id)
    paths = {
        "job_id": job_id,
        "job_dir": job_dir,
        "status": os.path.join(job_dir, "status.json"),
        "events": os.path.join(job_dir, "events.jsonl"),
        "run_log": os.path.join(job_dir, "run.log"),
        "stage_dir": os.path.join(job_dir, "stage"),
        "output_dir": os.path.join(job_dir, "outputs"),
        "raw_dir": os.path.join(job_dir, "raw"),
    }

    dir_keys = {"job_dir", "stage_dir", "output_dir", "raw_dir"}
    file_keys = {"status", "events", "run_log"}

    for k in dir_keys:
        os.makedirs(paths[k], exist_ok=True)

    for k in file_keys:
        os.makedirs(os.path.dirname(paths[k]), exist_ok=True)

    return paths
# def make_job_paths(base_dir, job_id,MA_DONVI):
#     job_dir = os.path.join(base_dir,MA_DONVI, job_id)

#     paths = {
#         "job_id": job_id,
#         "job_dir": job_dir,
#         "status": os.path.join(job_dir, "status.json"),
#         "events": os.path.join(job_dir, "events.jsonl"),
#         "run_log": os.path.join(job_dir, "run.log"),
#         "stage_dir": os.path.join(job_dir, "stage"),
#         "output_dir": os.path.join(job_dir, "outputs"),
#         "raw_dir": os.path.join(job_dir, "raw"),
#     }
#     # paths = {
#     #     "job_id": job_id,
#     #     "job_dir": job_dir,
#     #     "status": os.path.join(job_dir, "status.json"),
#     #     "events": os.path.join(job_dir, "events.jsonl"),
#     #     "stage_dir": os.path.join(job_dir, "stage"),
#     #     "output_dir": os.path.join(job_dir, "outputs"),
#     #     "raw_dir": os.path.join(job_dir, "raw"),
#     # }

#     for p in paths.values():
#         if isinstance(p, str) and (p.endswith(".json") or p.endswith(".jsonl")):
#             os.makedirs(os.path.dirname(p), exist_ok=True)
#         elif isinstance(p, str):
#             os.makedirs(p, exist_ok=True)

#     return paths

# =========================
# XML MAP + XML SAFE PARSE HELPERS
# =========================
def safe_str(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s:
        # Chuan hoa ve dang "dung san" (NFC). XML hoa don nhieu khi luu tieng Viet
        # o dang to hop (NFD), vi du "e" = "e" + dau mu + dau sac roi.
        # Bo chuyen UNICODE->TCVN3 ben VFP chi co bang anh xa dang dung san,
        # gap ky tu to hop se ra dau "?". Chuan hoa NFC o day giup tranh loi do.
        # NFC khong lam thay doi ASCII, so, duong dan -> an toan voi moi gia tri.
        s = unicodedata.normalize("NFC", s)
    return s

def norm_bool(v):
    s = safe_str(v).lower()
    return s in ("1", "true", "yes", "y", "x")

def xml_local_name(tag):
    if not tag:
        return ""
    return tag.split("}", 1)[-1] if "}" in tag else tag

def find_first_child_by_local_name(parent, name):
    if parent is None:
        return None
    for ch in list(parent):
        if xml_local_name(ch.tag) == name:
            return ch
    return None

def xml_find_node_by_path(root, path):
    """
    Tim node theo path day du:
      /HDon/DLHDon/TTChung/KHHDon
    hoac path tuong doi:
      THHDVu
      DVTinh
    Khong phu thuoc namespace.
    """
    if root is None:
        return None

    path = safe_str(path)
    if not path:
        return None

    parts = [p for p in path.strip("/").split("/") if p]
    if not parts:
        return None

    cur = root

    # Neu phan dau path trung ten root thi bo qua
    if parts and xml_local_name(cur.tag) == parts[0]:
        parts = parts[1:]

    for part in parts:
        cur = find_first_child_by_local_name(cur, part)
        if cur is None:
            return None

    return cur

def xml_find_text_by_path(root, path, default=""):
    node = xml_find_node_by_path(root, path)
    if node is None:
        return default
    txt = node.text
    if txt is None:
        return default
    return str(txt).strip()

def load_xml_map_from_excel(map_path):
    """
    Doc file XML map Excel.
    Yeu cau toi thieu:
      - node_name
      - target_table   (MASTER / LINE)
      - field_name
    """
    xls = pd.ExcelFile(map_path)
    if not xls.sheet_names:
        raise ValueError("File XML map khong co sheet nao")

    sheet = xls.sheet_names[0]
    df = pd.read_excel(map_path, sheet_name=sheet).fillna("")

    # Chuan hoa ten cot
    normalized_columns = []
    for c in df.columns:
        sc = safe_str(c)
        normalized_columns.append(sc)
    df.columns = normalized_columns

    required = ["node_name", "target_table", "field_name"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"XML map thieu cot bat buoc: {', '.join(missing)}")

    master_map = []
    line_map = []

    for _, row in df.iterrows():
        node_name = safe_str(row.get("node_name"))
        target_table = safe_str(row.get("target_table")).upper()
        field_name = safe_str(row.get("field_name"))

        if not node_name or not field_name or target_table not in ("MASTER", "LINE"):
            continue

        item = {
            "node_name": node_name,
            "target_table": target_table,
            "field_name": field_name,
        }

        if target_table == "MASTER":
            master_map.append(item)
        else:
            line_map.append(item)

    if not master_map and not line_map:
        raise ValueError("XML map khong co dong hop le nao")

    return {
        "master": master_map,
        "line": line_map,
    }
########################################
def normalize_invoice_text(v):
    s = safe_str(v)

    if not s:
        return ""

    # Bo khoang trang
    s = s.strip()

    # Neu la dang so co .0 thi cat ve so nguyen
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".", 1)[0]

    # Bo .0 o cuoi neu co
    if s.endswith(".0"):
        s = s[:-2]

    return s.strip().upper()

def make_invoice_key(huong, khhd, shd):
    return (
        normalize_invoice_text(huong),
        normalize_invoice_text(khhd),
        normalize_invoice_text(shd),
    )
############################################
def extract_thang_from_filename(file_path):
    """
    Lay thang tu ten file dang:
      HD_VAO_USA_MEVA_T3_KM.xlsx
      HD_RA_VINH_HOAN_T11.xlsx
      HD_VAO_XXX_T12_MTT.xlsx
    """
    base = os.path.basename(file_path).upper()
    m = re.search(r'_T(\d{1,2})(?:_|\.XLSX?$)', base)
    if m:
        return m.group(1)
    return ""
###############################################
# =========================
# LAY TANG DAN (--tang_dan)
# =========================
# Excel danh sach cua cong = "cong dang co gi". Hai cot ta tu them vao chinh file do
# = "ta da lam gi". Moi luot chay hop nhat hai thu ay thanh MOT file, nen khong bao
# gio phai giu hai ban roi lo chung lech nhau.
COT_DUONG_DAN = "Đường dẫn XML"
COT_KET_QUA = "Kết quả tải"

KQ_KHONG_GOC = "KHÔNG CÓ GỐC"                  # HTTP 500 - cong khong giu ban goc
KQ_MAT_TREN_CONG = "KHÔNG CÒN TRÊN CỔNG"  # luot truoc co, luot nay bien mat

# Lech o nhung cot NAY = hoa don doi ban chat -> phai tai lai.
# Co y KHONG so moi cot: cong chi can doi mot thu vat (viet hoa ten nguoi ban, doi
# kieu luu so) la toan bo hoa don bi coi la moi va tai lai tu dau - dung cai ta dang
# tranh, ma lai xay ra am tham. Da dinh dung bay nay mot lan khi chay thu.
COT_DOI_BAN_CHAT = [
    "Trạng thái hóa đơn",
    "Tổng tiền chưa thuế",
    "Tổng tiền thuế",
    "Tổng tiền chiết khấu thương mại",
    "Tổng tiền thanh toán",
]

_C_KHHD = "Ký hiệu hóa đơn"
_C_SOHD = "Số hóa đơn"


def _o(v):
    """Gia tri o -> chuoi chuan. Cong ghi o trong bang dau '-', phai coi nhu rong."""
    if v is None:
        return ""
    t = str(v).strip()
    return "" if t == "-" else t


def _bang_nhau(a, b):
    """So hai o. SO phai so nhu SO: cung mot con so nhung file tho tu cong luu kieu
    float ('214072722.0') con file da qua openpyxl luu kieu khac ('214072722')."""
    if a == b:
        return True
    try:
        return abs(float(a) - float(b)) < 0.01
    except (TypeError, ValueError):
        return a.strip().casefold() == b.strip().casefold()


def _ban_do_cot(ws, header_row=6):
    return {_o(ws.cell(header_row, c).value): c
            for c in range(1, ws.max_column + 1) if _o(ws.cell(header_row, c).value)}


def _bao_dam_cot(ws, ten, cot, header_row=6):
    """Chi so cot 'ten'; chua co thi them vao cuoi. Phai tim theo TEN truoc, khong thi
    moi luot hop nhat lai de them mot cot trung ten nua."""
    if ten in cot:
        return cot[ten]
    c = ws.max_column + 1
    ws.cell(header_row, c, ten)
    cot[ten] = c
    return c


def _khoa_dong(ws, r, cot):
    kh = _o(ws.cell(r, cot[_C_KHHD]).value)
    so = _o(ws.cell(r, cot[_C_SOHD]).value)
    return (kh, so) if kh and so else None


def _doc_sheet_tong(duong_dan, ten_sheet):
    """Đọc 1 sheet của Excel tổng thành list dict, GIỮ NGUYÊN kiểu ô.

    Cố tình KHÔNG dùng pd.read_excel: đọc dtype=str thì 51228 thành "51228.0",
    còn để pandas tự suy kiểu thì "00000003" thành số 3 — mất số 0 đầu, mà đó
    đúng là thứ BR-HD-01 vừa chuẩn hóa. openpyxl trả về đúng giá trị đã lưu.
    Cũng không dùng read_only=True (bài học: Excel thiếu thẻ <dimension> trả về
    lưới 1x1 mà không báo lỗi gì).
    """
    if not os.path.exists(duong_dan):
        return []
    wb = openpyxl.load_workbook(duong_dan, data_only=True)
    try:
        if ten_sheet not in wb.sheetnames:
            return []
        ws = wb[ten_sheet]
        dong = ws.iter_rows(values_only=True)
        try:
            tieu_de = next(dong)
        except StopIteration:
            return []
        cot = [str(c).strip() if c is not None else "" for c in tieu_de]
        return [{k: ("" if v is None else v) for k, v in zip(cot, r) if k}
                for r in dong]
    finally:
        wb.close()


def ghi_excel_tong(duong_dan, masters, lines, sheet_master, sheet_line, run_log):
    """Ghi Excel tổng theo kiểu HỢP NHẤT — KHÔNG ghi đè trắng.

    Vì sao (ca thật HOA_SANG T3 ngày 11/08): chế độ tăng dần bỏ qua hóa đơn đã
    tải, nên lượt chạy sau chỉ còn vài dòng đi qua hàng đợi. Ghi đè trắng bằng
    ngần ấy dòng làm Excel tổng từ 100 tụt xuống 29, mà bước nạp lại coi file
    này là TOÀN BỘ tháng. File vừa đóng vai "kết quả một lượt" vừa là "nguồn nạp
    DB" — hai vai xung khắc; hợp nhất là cách gỡ.

    Khóa gộp là MA_HD: hóa đơn có mặt trong lượt này thì bản MỚI thắng (kéo theo
    cả dòng hàng của nó), hóa đơn không đụng tới thì giữ nguyên.

    Hợp nhất hỏng thì TUYỆT ĐỐI không im lặng ghi đè: cất bản cũ ra tên khác đã.
    """
    thu_muc, ten = os.path.split(duong_dan)
    goc, duoi = os.path.splitext(ten)
    trang_thai = "moi"

    if os.path.exists(duong_dan):
        try:
            khoa_moi = {str(r.get("MA_HD", "")) for r in masters}
            cu_m = [r for r in _doc_sheet_tong(duong_dan, sheet_master)
                    if str(r.get("MA_HD", "")) not in khoa_moi]
            cu_l = [r for r in _doc_sheet_tong(duong_dan, sheet_line)
                    if str(r.get("MA_HD", "")) not in khoa_moi]
            trang_thai = f"hop_nhat giu_cu={len(cu_m)} luot_nay={len(masters)}"
            masters = cu_m + list(masters)
            lines = cu_l + list(lines)
        except Exception as e:
            sao_luu = os.path.join(thu_muc, f"~loi_hop_nhat_{goc}{duoi}")
            try:
                shutil.copy2(duong_dan, sao_luu)
            except OSError:
                pass
            append_run_log(run_log,
                f"EXCEL_TONG_HOP_NHAT_LOI {short_exc(e)} -> da sao luu {sao_luu}")
            trang_thai = "LOI_HOP_NHAT_da_sao_luu"

    # Ghi ra file tạm rồi đổi tên: .xlsx là file ZIP, chết giữa lúc ghi đè thẳng
    # sẽ để lại ZIP cụt không mở được. os.replace trên cùng ổ đĩa là nguyên tử.
    tam = os.path.join(thu_muc, f"~tam_{goc}{duoi}")
    with pd.ExcelWriter(tam) as writer:
        pd.DataFrame(masters).fillna("").to_excel(writer, sheet_name=sheet_master, index=False)
        pd.DataFrame(lines).fillna("").to_excel(writer, sheet_name=sheet_line, index=False)
    os.replace(tam, duong_dan)

    append_run_log(run_log,
        f"EXCEL_TONG_GHI {trang_thai} master={len(masters)} line={len(lines)} file={ten}")
    return len(masters), len(lines)


# ===== DỪNG ÊM =====
# Backend đặt file STOP vào job_dir để XIN dừng. Trước đây nút Dừng gọi thẳng
# proc.Kill(entireProcessTree) — giết cứng ngay giữa vòng tải, nên bước ghi cột
# "Đường dẫn XML" (chạy sau vòng tải) không kịp thực hiện: toàn bộ XML vừa tải
# về coi như mất dấu, lượt sau tải lại từ đầu. Kiểm cờ ở các mốc an toàn rồi
# thoát theo đường bình thường thì bước đó vẫn chạy.
TEN_FILE_DUNG = "STOP"
_LUC_BAT_DAU = time.time()

def nen_dung(job_dir):
    """True nếu backend đã xin dừng.

    So theo thời điểm sửa file: STOP sót lại từ lần chạy TRƯỚC phải bị bỏ qua,
    không thì mọi lần chạy sau đều tự dừng ngay khi vừa khởi động. Trừ hao 5
    giây cho sai lệch đồng hồ / độ phân giải mtime của hệ thống tệp.
    """
    try:
        return os.path.getmtime(os.path.join(job_dir, TEN_FILE_DUNG)) >= _LUC_BAT_DAU - 5
    except OSError:
        return False   # không có file, hoặc không đọc được -> cứ chạy tiếp


def hop_nhat_excel(duong_dan_cu, noi_dung_moi, thang, job_dir, run_log, header_row=6):
    """Hop nhat Excel VUA TAI (bytes) voi file dang co, ghi de len chinh ten cu.

    Ban MOI lam goc (no la su that ve cong HIEN GIO), roi chep hai cot chu thich cua
    ban cu sang. Tra ve dict thong ke, hoac None neu khong hop nhat duoc - khi do
    nguoi goi ghi thang bytes va chay day du nhu cu.
    """
    import shutil
    # File tam PHAI co duoi .xlsx (openpyxl tu choi duoi la, khong doc duoc), nhung
    # tien to '~tam_' de no khong lot vao glob("HD_*.xlsx") neu chuong trinh chet giua
    # chung va bo lai rac.
    def _tam(nhan):
        return os.path.join(os.path.dirname(duong_dan_cu),
                            "~tam_" + nhan + "_" + os.path.basename(duong_dan_cu))

    tam_moi = _tam("moi")
    with open(tam_moi, "wb") as f:
        f.write(noi_dung_moi)

    try:
        # KHONG dung read_only: Excel THO tu cong khong khai the <dimension>, o che do
        # read_only openpyxl bao bang chi co 1 dong 1 cot - doc ra rong ma khong bao loi.
        wb_cu = openpyxl.load_workbook(duong_dan_cu, data_only=True)
        ws_cu = wb_cu.active
        cot_cu = _ban_do_cot(ws_cu, header_row)
        if _C_KHHD not in cot_cu or _C_SOHD not in cot_cu:
            raise ValueError("file cu thieu cot khoa")

        cu, trung = {}, 0
        for r in range(header_row + 1, ws_cu.max_row + 1):
            k = _khoa_dong(ws_cu, r, cot_cu)
            if not k:
                continue
            if k in cu:
                trung += 1
            cu[k] = {t: _o(ws_cu.cell(r, c).value) for t, c in cot_cu.items()}
        wb_cu.close()

        # Trung khoa = hai hoa don khac nhau doi chung mot ten. Hop nhat luc nay la ghi
        # chu thich SAI VINH VIEN vao file. Tha chay cham nhu cu con hon.
        if trung:
            append_run_log(run_log, f"HOP_NHAT_BO_QUA trung_khoa={trung} -> chay day du")
            return None

        wb_moi = openpyxl.load_workbook(tam_moi)
        ws_moi = wb_moi.active
        cot_moi = _ban_do_cot(ws_moi, header_row)
        if _C_KHHD not in cot_moi or _C_SOHD not in cot_moi:
            raise ValueError("file moi thieu cot khoa - cong da doi cau truc?")

        c_dd = _bao_dam_cot(ws_moi, COT_DUONG_DAN, cot_moi, header_row)
        c_kq = _bao_dam_cot(ws_moi, COT_KET_QUA, cot_moi, header_row)

        giu, doi_ban_chat, thay_doi, con_o_moi = 0, 0, [], set()

        for r in range(header_row + 1, ws_moi.max_row + 1):
            k = _khoa_dong(ws_moi, r, cot_moi)
            if not k:
                continue
            con_o_moi.add(k)
            c = cu.get(k)
            if not c:
                continue      # hoa don MOI - de trong chu thich, se tai

            khac = [t for t in COT_DOI_BAN_CHAT
                    if t in cot_moi and t in c
                    and not _bang_nhau(_o(ws_moi.cell(r, cot_moi[t]).value), c[t])]
            if khac:
                # Doi ban chat -> KHONG chep chu thich sang, de no duoc tai lai
                doi_ban_chat += 1
                thay_doi.append((k, [(t, c[t], _o(ws_moi.cell(r, cot_moi[t]).value))
                                     for t in khac]))
                continue

            ws_moi.cell(r, c_dd, c.get(COT_DUONG_DAN, ""))
            ws_moi.cell(r, c_kq, c.get(COT_KET_QUA, ""))
            if c.get(COT_DUONG_DAN) or c.get(COT_KET_QUA) == KQ_KHONG_GOC:
                giu += 1

        # Dong chi con o ban cu: GIU LAI, khong xoa. Nguyen nhan pho bien khong phai
        # cong xoa hoa don, ma la KHOANG NGAY hai luot khac nhau - xoa di la mat sach
        # chu thich cua ca thang, roi luot sau tai lai tu dau, am tham.
        mat = 0
        for k, c in cu.items():
            if k in con_o_moi:
                continue
            mat += 1
            r = ws_moi.max_row + 1
            for t, v in c.items():
                if t in cot_moi:
                    ws_moi.cell(r, cot_moi[t], v)
            ws_moi.cell(r, c_kq, c.get(COT_KET_QUA) or KQ_MAT_TREN_CONG)

        tam_ghi = _tam("ghi")
        wb_moi.save(tam_ghi)
        wb_moi.close()
        os.replace(tam_ghi, duong_dan_cu)

        if thay_doi:
            # Ghi lai NHUNG KHONG dong vao database. Cau hoi "hoa don da nap bi dieu
            # chinh thi xu ly ban cu the nao" con dang cho ke toan tra loi; khong ghi
            # bay gio thi toi luc co cau tra loi, lich su da troi mat.
            try:
                with open(os.path.join(job_dir, f"THAY_DOI_T{thang}.txt"), "a",
                          encoding="utf-8") as f:
                    f.write("--- " + datetime.datetime.now().isoformat(timespec="seconds")
                            + " | " + os.path.basename(duong_dan_cu) + "\n")
                    for k, cac in thay_doi:
                        f.write("  " + k[0] + "/" + k[1] + "\n")
                        for t, cu_v, moi_v in cac:
                            f.write("      " + t + ": '" + cu_v + "' -> '" + moi_v + "'\n")
            except Exception as e:
                append_run_log(run_log, f"GHI_THAY_DOI_LOI {short_exc(e)}")

            # Giu ban tai THO cua nhung ngay CO thay doi, de sau con doi chieu duoc
            # "hom do cong noi gi". Ngay khong thay doi thi bo, khoi phinh o dia.
            try:
                ls = os.path.join(job_dir, "lichsu")
                os.makedirs(ls, exist_ok=True)
                dau = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                shutil.copy2(tam_moi, os.path.join(
                    ls, os.path.basename(duong_dan_cu)[:-5] + "_" + dau + ".xlsx"))
            except Exception as e:
                append_run_log(run_log, f"LUU_LICH_SU_LOI {short_exc(e)}")

        return {"giu": giu, "doi_ban_chat": doi_ban_chat, "mat": mat,
                "tong": len(con_o_moi)}

    except Exception as e:
        append_run_log(run_log, f"HOP_NHAT_LOI {short_exc(e)} -> chay day du")
        return None
    finally:
        try:
            os.remove(tam_moi)
        except OSError:
            pass


def loc_hoa_don_can_tai(ds_hd, excel_path, run_log, header_row=6):
    """Chia ds_hd thanh (can_tai, dung_lai_file_cu, so_bo_qua, so_khong_goc).

    'dung_lai_file_cu' = hoa don KHONG can tai nhung file XML VAN CON tren dia. Bat
    buoc phai tra ve: Excel TONG duoc dung LAI TU DAU moi luot chay, chi tu nhung XML
    di qua hang doi chuyen doi. Bo tai ma khong day file cu vao hang doi thi hoa don
    bien mat khoi Excel tong -> buoc nap khong nhin thay -> file nam ly lai trong raw\\.
    Do dung la thu da xay ra: bo qua 126 hoa don, Excel tong tut tu 126 xuong 27 dong.

    Cai dat tien la KHONG GOI MANG (moi request con kem sleep 1 giay). Doc lai file
    XML co san tren dia thi nhanh, khong dang de danh doi tinh day du cua Excel tong.

    File khong con tren dia = hoa don DA NAP THANH CONG tu truoc: MoveArtifacts chi
    chay sau khi nap xong, nen file bien mat khoi raw\\ chinh la bang chung da vao so.
    Khong tai lai, khong dua vao Excel tong — no da nam trong database roi.

    Doc hong thi tra ve NGUYEN ds_hd. Sai lam te nhat o day la BO SOT hoa don, khong
    phai tai thua.
    """
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        cot = _ban_do_cot(ws, header_row)
        if COT_DUONG_DAN not in cot and COT_KET_QUA not in cot:
            wb.close()
            return ds_hd, [], 0, 0

        duong_dan, khong_goc = {}, set()
        for r in range(header_row + 1, ws.max_row + 1):
            k = _khoa_dong(ws, r, cot)
            if not k:
                continue
            if COT_DUONG_DAN in cot:
                dd = _o(ws.cell(r, cot[COT_DUONG_DAN]).value)
                if dd:
                    duong_dan[k] = dd
            if COT_KET_QUA in cot and _o(ws.cell(r, cot[COT_KET_QUA]).value) == KQ_KHONG_GOC:
                khong_goc.add(k)
        wb.close()

        can_tai, dung_lai, da_nap_truoc = [], [], 0
        for h in ds_hd:
            k = (str(h["khhdon"]).strip(), str(h["shdon"]).strip())
            if k in khong_goc:
                continue                      # ket qua cuoi cung, dung thu lai
            dd = duong_dan.get(k)
            if not dd:
                can_tai.append(h)
            elif os.path.exists(dd):
                dung_lai.append((h, dd))      # con file -> van phai dua vao Excel tong
            else:
                # CO duong dan nhung MAT file = DA NAP XONG tu luot truoc.
                # Suy luan nay dung vi MoveArtifacts CHI chay sau khi nap THANH CONG —
                # file bien mat khoi raw\ chinh la bang chung da vao so. (Chot voi
                # Truong 11/08: cot 'Duong dan XML' du tin cay, chap nhan rui ro con
                # lai la ai do xoa file bang tay.)
                da_nap_truoc += 1

        if da_nap_truoc:
            append_run_log(run_log,
                f"TANG_DAN_DA_NAP {da_nap_truoc} hoa don da nap tu truoc (file da bi doi di)")
        return can_tai, dung_lai, len(ds_hd) - len(can_tai), len(khong_goc)
    except Exception as e:
        append_run_log(run_log, f"LOC_CAN_TAI_LOI {short_exc(e)} -> tai tat")
        return ds_hd, [], 0, 0


def them_cot_xml_path(excel_path, xml_paths, header_row=6, col_khhd="Ký hiệu hóa đơn", col_shd="Số hóa đơn", col_moi="Đường dẫn XML", ket_qua_tai=None):
    """
    Mo file Excel tai ve tu server, dien cot 'Duong dan XML' (va 'Ket qua tai'),
    khop theo (khhd, shd).

    Hai thay doi so voi ban dau, deu bat buoc cho che do tang dan:
      - Tim cot theo TEN truoc khi them moi. Ban cu luon them vao cuoi, nen file da
        hop nhat se moc them mot cot trung ten sau moi luot chay.
      - Chi ghi de khi CO gia tri moi. Ban cu ghi xml_paths.get(key, "") cho moi dong,
        tuc la xoa trang duong dan cua nhung hoa don da bi bo qua o luot nay - luot
        sau lai tai lai tu dau, dung cai ta dang tranh.
    """
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # Tim cot khhd va shd trong header_row
        idx_khhd = None
        idx_shd = None
        max_col = ws.max_column
        for c in range(1, max_col + 1):
            val = ws.cell(row=header_row, column=c).value
            if val is None:
                continue
            val_str = str(val).strip()
            if val_str == col_khhd:
                idx_khhd = c
            elif val_str == col_shd:
                idx_shd = c

        if idx_khhd is None or idx_shd is None:
            #print(f"      [Excel] Khong tim thay cot '{col_khhd}' hoac '{col_shd}' o dong {header_row}")
            wb.close()
            return False

        cot_hien_co = _ban_do_cot(ws, header_row)
        col_xml = _bao_dam_cot(ws, col_moi, cot_hien_co, header_row)
        col_kq = _bao_dam_cot(ws, COT_KET_QUA, cot_hien_co, header_row)

        # Duyet tung dong data, khop va dien path
        count_match = 0
        for r in range(header_row + 1, ws.max_row + 1):
            khhd = ws.cell(row=r, column=idx_khhd).value
            shd = ws.cell(row=r, column=idx_shd).value
            if khhd is None or shd is None:
                continue
            key = (str(khhd).strip(), str(shd).strip())
            # CHI ghi de khi co gia tri moi — dong bi bo qua o luot nay phai giu nguyen
            # chu thich cu, khong thi luot sau tai lai tu dau.
            xml_path = xml_paths.get(key)
            if xml_path:
                ws.cell(row=r, column=col_xml, value=xml_path)
                count_match += 1
            if ket_qua_tai and key in ket_qua_tai:
                ws.cell(row=r, column=col_kq, value=ket_qua_tai[key])

        wb.save(excel_path)
        wb.close()
        #print(f"      [Excel] Da them cot 'Duong dan XML' ({count_match} hoa don co XML)")
        return True
    except Exception as e:
        #print(f"      [Excel] Loi them cot: {e}")
        return False

def an_toan_du_lieu(text, max_bytes=254):
    """
    Chuan bi gia tri ghi vao DBF codepage utf8 (giu nguyen tieng Viet co dau).
    Chi cat theo so byte UTF-8 cho khop field C(max_bytes), khong cat giua
    1 ky tu nhieu byte (dung errors='ignore' khi decode lai).
    """
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)
    text = text.strip()
    b = text.encode("utf-8")[:max_bytes]
    return b.decode("utf-8", "ignore")

# def excel_to_exact_dbf(excel_path, output_dir):
#     """
#     Chuyen tung sheet trong Excel sang DBF.
#     Ten field DBF lay theo header Excel.
#     Vi field_name trong map da duoc quy chuan DBF, ham nay se:
#       - giu toi da ten field goc
#       - chi chuan hoa nhe de tranh loi
#       - dam bao khong trung ten field
#     """
#     def normalize_dbf_field_name(name, used_names):
#         name = safe_str(name).upper()

#         # Bo dau neu co
#         name = unicodedata.normalize('NFD', name)
#         name = ''.join(c for c in name if unicodedata.category(c) != 'Mn')
#         name = name.replace('Đ', 'D').replace('đ', 'd')

#         # Chi giu A-Z 0-9 _
#         name = re.sub(r'[^A-Z0-9_]', '_', name)
#         name = re.sub(r'_+', '_', name).strip('_')

#         if not name:
#             name = "FIELD"

#         # DBF an toan <= 10 ky tu
#         base = name[:10]
#         candidate = base
#         i = 1

#         while candidate in used_names:
#             suffix = str(i)
#             candidate = base[:10 - len(suffix)] + suffix
#             i += 1

#         used_names.add(candidate)
#         return candidate

#     try:
#         xl = pd.ExcelFile(excel_path)
#         sheet_names = xl.sheet_names
#         #print(f"---sheet processing... : {sheet_names} ---")

#         for sheet in sheet_names:
#             df = pd.read_excel(excel_path, sheet_name=sheet)
#             df = df.fillna('')

#             dbf_filename = sheet.upper() + ".DBF"
#             dbf_path = os.path.join(output_dir, dbf_filename)

#             original_cols = list(df.columns)
#             used_names = set()
#             final_cols = [normalize_dbf_field_name(col, used_names) for col in original_cols]
#             df.columns = final_cols

#             specs = [f"{col} C(254)" for col in df.columns]
#             table_spec = "; ".join(specs)

#             # Xoa file DBF cu neu ton tai
#             if os.path.exists(dbf_path):
#                 try:
#                     os.remove(dbf_path)
#                 except Exception:
#                     pass

#             # codepage utf8 (byte header 0xf0) de giu tieng Viet co dau day du nhu Excel.
#             # Luu y: app doc DBF (vd VFP) phai ho tro doc UTF-8 thi moi hien dung.
#             table = dbf.Table(dbf_path, table_spec, codepage='utf8', dbf_type='vfp')
#             table.open(mode=dbf.READ_WRITE)

#             for _, row in df.iterrows():
#                 data = tuple(an_toan_du_lieu(str(val).strip()) for val in row)
#                 table.append(data)

#             table.close()
#             #print(f"Done: {dbf_filename}")
    
#     except Exception as e:
#         #print(f"System error: {repr(e)}")
#         raise
#     finally:
#         table.close()   # luôn chạy, kể cả khi append lỗi

# =========================
# EXCEL FALLBACK + CAP NHAT TTHAI_HD TU EXCEL TONG HOP
# =========================
def normalize_header_text(v):
    s = safe_str(v)
    s = re.sub(r'\s+', ' ', s).strip().lower()
    return s

def build_header_map(df):
    out = {}
    for c in df.columns:
        out[normalize_header_text(c)] = c
    return out

def pick_col(header_map, candidates):
    for cand in candidates:
        key = normalize_header_text(cand)
        if key in header_map:
            return header_map[key]
    return None

def value_from_row(row, col_name):
    if not col_name:
        return ""
    try:
        v = row[col_name]
        if pd.isna(v):
            return ""
        s = str(v).strip()
        if s:
            # Chuan hoa ve dang "dung san" (NFC) - xem ghi chu o safe_str().
            # File Excel tong hop tai tu cong Thue cung co the o dang to hop (NFD).
            s = unicodedata.normalize("NFC", s)
        return s
    except Exception:
        return ""

def read_excel_summary_rows(excel_path):
    """
    Doc file Excel tong hop GDT, header o dong 6.
    Tra ve DataFrame.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    header_row = 6
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [safe_str(h) for h in headers]

    data = []
    for r in range(header_row + 1, ws.max_row + 1):
        row_dict = {}
        non_empty = 0

        for c, h in enumerate(headers, start=1):
            val = ws.cell(row=r, column=c).value
            sval = "" if val is None else str(val).strip()
            row_dict[h] = sval
            if sval != "":
                non_empty += 1

        if non_empty == 0:
            continue

        data.append(row_dict)

    wb.close()

    if not data:
        return pd.DataFrame()

    return pd.DataFrame(data).fillna("")

def update_master_tthai_hd_from_excel_files(excel_file_infos, all_master_rows, lock_rows, run_log=None):
    """
    Cap nhat TTHAI_HD cho TAT CA cac hoa don da co trong all_master_rows
    bang cach doi chieu voi cac file Excel tong hop.
    Ap dung cho ca VAO va RA.
    Match theo 2 cap khoa:
      1) HUONG + KHHD + SO_HD
      2) HUONG + KHHD_TRA_C + SHD_TRA_CU
    """
    index_map = {}

    with lock_rows:
        for row in all_master_rows:
            huong = safe_str(row.get("HUONG")).upper()

            key1 = make_invoice_key(
                huong,
                row.get("KHHD"),
                row.get("SO_HD"),
            )
            key2 = make_invoice_key(
                huong,
                row.get("KHHD_TRA_C"),
                row.get("SHD_TRA_CU"),
            )

            if key1[1] and key1[2]:
                index_map[key1] = row

            if key2[1] and key2[2]:
                index_map[key2] = row

    updated = 0

    for info in excel_file_infos:
        excel_path = info.get("path")
        huong = safe_str(info.get("huong")).upper()

        if not excel_path or not os.path.exists(excel_path):
            continue

        try:
            df = read_excel_summary_rows(excel_path)
            if df.empty:
                continue

            header_map = build_header_map(df)

            col_khhd = pick_col(header_map, ["Ký hiệu hóa đơn", "Ky hieu hoa don"])
            col_shd = pick_col(header_map, ["Số hóa đơn", "So hoa don"])
            col_tthai_hd = pick_col(header_map, ["Trạng thái hóa đơn", "Trang thai hoa don"])

            if not col_khhd or not col_shd or not col_tthai_hd:
                if run_log:
                    append_run_log(run_log, f"TTHAI_HD_SKIP file={excel_path} missing_required_columns")
                continue

            updated_this_file = 0

            for _, row_excel in df.iterrows():
                khhd = value_from_row(row_excel, col_khhd)
                shd = value_from_row(row_excel, col_shd)
                tthai_hd_unicode = value_from_row(row_excel, col_tthai_hd)

                key = make_invoice_key(huong, khhd, shd)
                row_master = index_map.get(key)

                if row_master is not None:
                    row_master["TTHAI_HD"] = safe_str(tthai_hd_unicode)
                    updated += 1
                    updated_this_file += 1

            if run_log:
                append_run_log(run_log, f"TTHAI_HD_UPDATED file={excel_path} matched={updated_this_file}")

        except Exception as e:
            if run_log:
                append_run_log(run_log, f"TTHAI_HD_ERROR file={excel_path} err={short_exc(e)}")

    return updated

def build_rows_from_excel_without_xml(excel_path, huong_default="", ma_donvi="", mst_tra_cuu="", thang="", nam=""):
    """
    Chi dung cho file Excel tong hop HOA DON VAO KHONG MA (_KM.xlsx).
    Lay cac dong KHONG CO duong dan XML va sinh:
      - 1 dong master
      - 1 dong line dai dien

    Quy tac line:
      - SO_LUONG = 1
      - DON_GIA = TONG_TIEN
      - THANH_TIEN = TONG_TIEN
      - TEN_HANG = "NOXML - " + TEN_BAN
      - DVT = "Lần"

    TTHAI_HD lay tu Excel (Unicode).
    """
    master_rows = []
    line_rows = []

    base_name = os.path.basename(excel_path).upper()
    if "_KM" not in base_name:
        return master_rows, line_rows

    df = read_excel_summary_rows(excel_path)
    if df.empty:
        return master_rows, line_rows

    header_map = build_header_map(df)

    col_xml = pick_col(header_map, ["Đường dẫn XML", "Duong dan XML"])
    col_kieu_hd = pick_col(header_map, ["Ký hiệu mẫu số", "Ky hieu mau so"])
    col_khhd = pick_col(header_map, ["Ký hiệu hóa đơn", "Ky hieu hoa don"])
    col_shd = pick_col(header_map, ["Số hóa đơn", "So hoa don"])
    col_ngay_hd = pick_col(header_map, ["Ngày lập", "Ngay lap"])

    col_mst_ban = pick_col(header_map, [
        "MST người bán/MST người xuất hàng",
        "Ma so thue nguoi ban/MST nguoi xuat hang"
    ])
    col_ten_ban = pick_col(header_map, [
        "Tên người bán/Tên người xuất hàng",
        "Ten nguoi ban/Ten nguoi xuat hang"
    ])

    col_mst_mua = pick_col(header_map, [
        "MST người mua/MST người nhận hàng",
        "Ma so thue nguoi mua/MST nguoi nhan hang"
    ])
    col_ten_mua = pick_col(header_map, [
        "Tên người mua/Tên người nhận hàng",
        "Ten nguoi mua/Ten nguoi nhan hang"
    ])
    col_dchi_mua = pick_col(header_map, [
        "Địa chỉ người mua",
        "Dia chi nguoi mua"
    ])

    col_tien_hang = pick_col(header_map, [
        "Tổng tiền chưa thuế",
        "Tong tien chua thue"
    ])
    col_tien_vat = pick_col(header_map, [
        "Tổng tiền thuế",
        "Tong tien thue"
    ])
    col_tien_ck = pick_col(header_map, [
        "Tổng tiền chiết khấu thương mại",
        "Tong tien chiet khau thuong mai"
    ])
    col_tien_phi = pick_col(header_map, [
        "Tổng tiền phí",
        "Tong tien phi"
    ])
    col_tong_tien = pick_col(header_map, [
        "Tổng tiền thanh toán",
        "Tong tien thanh toan"
    ])
    col_dvt_te = pick_col(header_map, [
        "Đơn vị tiền tệ",
        "Don vi tien te"
    ])
    col_ty_gia = pick_col(header_map, [
        "Tỷ giá",
        "Ty gia"
    ])
    col_tthai_hd = pick_col(header_map, [
        "Trạng thái hóa đơn",
        "Trang thai hoa don"
    ])
    col_kq_kiemtra = pick_col(header_map, [
        "Kết quả kiểm tra hóa đơn",
        "Ket qua kiem tra hoa don"
    ])

    huong_norm = normalize_invoice_text(huong_default)
    ma_donvi_norm = safe_str(ma_donvi).strip().upper()
    mst_norm = normalize_invoice_text(mst_tra_cuu)
    thang_norm = normalize_invoice_text(thang)
    nam_norm = normalize_invoice_text(nam)

    for _, row in df.iterrows():
        xml_path = value_from_row(row, col_xml)

        if xml_path:
            continue

        kieu_hd = value_from_row(row, col_kieu_hd)
        khhd = normalize_invoice_text(value_from_row(row, col_khhd))
        shd = normalize_invoice_text(value_from_row(row, col_shd))
        ngay_hd = value_from_row(row, col_ngay_hd)

        mst_ban = value_from_row(row, col_mst_ban)
        ten_ban = value_from_row(row, col_ten_ban)

        mst_mua = value_from_row(row, col_mst_mua)
        ten_mua = value_from_row(row, col_ten_mua)
        dchi_mua = value_from_row(row, col_dchi_mua)

        tien_hang = value_from_row(row, col_tien_hang)
        tien_vat = value_from_row(row, col_tien_vat)
        tien_ck = value_from_row(row, col_tien_ck)
        tien_phi = value_from_row(row, col_tien_phi)
        tong_tien = value_from_row(row, col_tong_tien)
        dvt_te = value_from_row(row, col_dvt_te)
        ty_gia = value_from_row(row, col_ty_gia)
        tthai_hd = value_from_row(row, col_tthai_hd)
        kq_kiemtra = value_from_row(row, col_kq_kiemtra)

        # ma_hd = f"{ma_donvi_norm}_{mst_norm}_{khhd}_{shd}"
        # MỚI: mst người bán ở nhánh này lấy từ biến MST người bán có sẵn trong cùng hàm
        #      (tên kiểu mst_nban / nbmst — nhìn 15 dòng phía trên); fallback mst_norm
        mst_ph = str(mst_ban or mst_norm).strip()
        ma_hd = f"{huong_norm}_{mst_ph}_{khhd}_{shd}"
        # ma_hd = f"{huong_norm}_{(mst_ban or mst_norm)}_{khhd}_{shd}"
        master_row = {
            "MA_HD": ma_hd,
            "MA_DONVI": ma_donvi_norm,
            "MST_TRA_CU": mst_norm,
            "HUONG": huong_norm,
            "THANG": thang_norm,
            "NAM": nam_norm,
            "KHHD_TRA_C": khhd,
            "SHD_TRA_CU": shd,
            "XML_PATH": "",
            "NGUON_DL": "EXCEL_NO_XML",

            "KIEU_HD": kieu_hd,
            "KHHD": khhd,
            "SO_HD": shd,
            "NGAY_HD": ngay_hd,

            "MST_BAN": mst_ban,
            "TEN_BAN": ten_ban,

            "MST_MUA": mst_mua,
            "TEN_MUA": ten_mua,
            "DCHI_MUA": dchi_mua,

            "TIEN_HANG": tien_hang,
            "TIEN_VAT": tien_vat,
            "TIEN_CK": tien_ck,
            "TIEN_PHI": tien_phi,
            "TONG_TIEN": tong_tien,
            "DVT_TE": dvt_te,
            "TY_GIA": ty_gia,

            "TTHAI_HD": tthai_hd,
            "KQ_KTRA": kq_kiemtra,
        }

        line_row = {
            "MA_HD": ma_hd,
            "HUONG": huong_norm,
            "THANG": thang_norm,
            "NAM": nam_norm,
            "LINE_NO": "1",
            "NGUON_DL": "EXCEL_NO_XML",

            "TEN_HANG": f"NOXML - {ten_ban}" if ten_ban else "NOXML",
            "SO_LUONG": "1",
            "DVT": "Lần",
            "DON_GIA": tong_tien,
            "THANH_TIEN": tong_tien,
            "PT_VAT": "",
        }

        master_rows.append(master_row)
        line_rows.append(line_row)

    return master_rows, line_rows

def append_non_xml_rows_from_excel_files(
    excel_file_infos,
    all_master_rows,
    all_line_rows,
    lock_rows,
    ma_donvi="",
    mst_tra_cuu="",
    nam="",
    run_log=None
):
    """
    Chi bo sung tu file HOA DON VAO KHONG MA (_KM.xlsx).
    Lay cac dong khong co XML roi append vao all_master_rows / all_line_rows.
    Tu dong lay THANG tu ten file.
    """
    added_master = 0
    added_line = 0

    for info in excel_file_infos:
        excel_path = info.get("path")
        huong = info.get("huong", "")

        if not excel_path or not os.path.exists(excel_path):
            continue

        base_name = os.path.basename(excel_path).upper()
        if huong != "VAO" or "_KM" not in base_name:
            continue

        try:
            thang_file = extract_thang_from_filename(excel_path)

            master_rows, line_rows = build_rows_from_excel_without_xml(
                excel_path=excel_path,
                huong_default=huong,
                ma_donvi=ma_donvi,
                mst_tra_cuu=mst_tra_cuu,
                thang=thang_file,
                nam=nam
            )

            if master_rows or line_rows:
                # Bỏ hóa đơn ĐÃ có trong danh sách. Trước đây extend thẳng, không
                # kiểm gì: hóa đơn nào vừa đi qua hàng đợi XML vừa xuất hiện ở đây
                # sẽ nằm HAI dòng trong Excel tổng, và bước nạp ghi đè lẫn nhau.
                # Bản đọc từ XML luôn đầy đủ hơn nên nó được giữ, bản NOXML bị loại.
                with lock_rows:
                    da_co = {str(r.get("MA_HD", "")) for r in all_master_rows}
                    master_rows = [r for r in master_rows
                                   if str(r.get("MA_HD", "")) not in da_co]
                    khoa_them = {str(r.get("MA_HD", "")) for r in master_rows}
                    line_rows = [r for r in line_rows
                                 if str(r.get("MA_HD", "")) in khoa_them]
                    all_master_rows.extend(master_rows)
                    all_line_rows.extend(line_rows)

                added_master += len(master_rows)
                added_line += len(line_rows)

                if run_log:
                    append_run_log(
                        run_log,
                        f"APPEND_NO_XML_FROM_EXCEL file={excel_path} huong={huong} thang={thang_file} master={len(master_rows)} line={len(line_rows)}"
                    )
        except Exception as e:
            if run_log:
                append_run_log(run_log, f"APPEND_NO_XML_FROM_EXCEL_ERROR file={excel_path} err={short_exc(e)}")

    return added_master, added_line


def parse_xml_hoa_don(xml_path, ma_hd, MA_DONVI, mst_value, huong, thang, nam, khhd, shd, xml_map):
    """
    Doc 1 file XML hoa don theo map Excel.
    - MASTER: lay 1 lan tren toan bo XML
    - LINE: lap theo tung node HHDVu
    - Neu node khong ton tai -> de rong, khong bao loi
    """
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()

        # ===== Row master co san 1 so field he thong =====
        master_row = {
            "MA_HD": safe_str(ma_hd),
            "MA_DONVI": safe_str(MA_DONVI),
            "MST_TRA_CU": safe_str(mst_value),
            "HUONG": safe_str(huong),
            "THANG": safe_str(thang),
            "NAM": safe_str(nam),
            "KHHD_TRA_C": safe_str(khhd),
            "SHD_TRA_CU": safe_str(shd),
            "XML_PATH": safe_str(xml_path),
            "TTHAI_HD": "",
        }        
        
        # ===== Lay field MASTER tu map =====
        for item in xml_map.get("master", []):
            field_name = safe_str(item.get("field_name"))
            node_path = safe_str(item.get("node_name"))

            val = xml_find_text_by_path(root, node_path, default="")
            master_row[field_name] = safe_str(val)

        # ===== Tim tat ca HHDVu =====
        hhdvu_nodes = []
        for node in root.iter():
            if xml_local_name(node.tag) == "HHDVu":
                hhdvu_nodes.append(node)

        line_rows = []

        for idx, hhdvu in enumerate(hhdvu_nodes, 1):
            line_row = {
                "MA_HD": safe_str(ma_hd),
                "HUONG": safe_str(huong),
                "THANG": safe_str(thang),
                "NAM": safe_str(nam),
                "LINE_NO": str(idx),
            }

            for item in xml_map.get("line", []):
                field_name = safe_str(item.get("field_name"))
                node_path = safe_str(item.get("node_name"))

                # Cat path sau HHDVu de tim tuong doi tren chinh node line
                parts = [p for p in node_path.strip("/").split("/") if p]
                if "HHDVu" in parts:
                    pos = parts.index("HHDVu")
                    rel_parts = parts[pos + 1:]
                    rel_path = "/".join(rel_parts)
                else:
                    rel_path = node_path

                val = xml_find_text_by_path(hhdvu, rel_path, default="")
                line_row[field_name] = safe_str(val)

            line_rows.append(line_row)

        return True, master_row, line_rows

    except Exception as e:
        return False, None, str(e)
    
# def parse_xml_hoa_don(xml_path, ma_hd, MA_DONVI, mst_value, huong, thang, nam, khhd, shd):
#     """Doc 1 file XML hoa don, tra ve (master_row, line_rows)"""
#     try:
#         tree = ET.parse(xml_path)
#         root = tree.getroot()
#         nban = root.find('.//NBan')
#         nmua = root.find('.//NMua')
#         tt_chung = root.find('.//TTChung')
#         # Lay tong tien tu TToan
#         ttoan = root.find('.//TToan')
#         tg_cthue = ttoan.findtext('TgTCThue') if ttoan is not None else None
#         tg_thue = ttoan.findtext('TgTThue') if ttoan is not None else None

#         master_row = {
#             "ma_hd": ma_hd,
#             "KIEU_HD": tt_chung.findtext('.//KHMSHDon') if tt_chung is not None else "",
#             "KHHD": to_tcvn3(tt_chung.findtext('.//KHHDon')) if tt_chung is not None else "",
#             "SO_HD": to_tcvn3(tt_chung.findtext('.//SHDon')) if tt_chung is not None else "",
#             "NGAY_HD": tt_chung.findtext('.//NLap') if tt_chung is not None else "",
#             "TEN_BAN": to_tcvn3(nban.findtext('Ten')) if nban is not None else "",
#             "MST_BAN": nban.findtext('MST') if nban is not None else "",
#             "DIA_CHI_BAN": to_tcvn3(nban.findtext('DChi')) if nban is not None else "",
#             "TEN_MUA": to_tcvn3(nmua.findtext('Ten')) if nmua is not None else "",
#             "MST_MUA": nmua.findtext('MST') if nmua is not None else "",
#             "DIA_CHI_MUA": to_tcvn3(nmua.findtext('DChi')) if nmua is not None else "",
#             "NG_GD": to_tcvn3(nmua.findtext('HVTNMHang')) if nmua is not None else "",
#             "TIEN_HANG": tg_cthue,
#             "TIEN_VAT": tg_thue,
#             "TIEN_CK": ttoan.findtext('TTCKTMai') if ttoan is not None else "",
#         }

#         line_rows = []
#         for hhdvu in root.findall('.//HHDVu'):
#             line_rows.append({
#                 "ma_hd": ma_hd,
#                 "LOAI_HH": to_tcvn3(hhdvu.findtext('TChat')) if hhdvu is not None else "",
#                 "STT_LINE": hhdvu.findtext('STT'),
#                 "MA_NGAN": to_tcvn3(hhdvu.findtext('MHHDVu')) if hhdvu is not None else "",
#                 "TEN_HANG": to_tcvn3(hhdvu.findtext('THHDVu')) if hhdvu is not None else "",
#                 "DVT": to_tcvn3(hhdvu.findtext('DVTinh')) if hhdvu is not None else "",
#                 "SO_LUONG": hhdvu.findtext('SLuong'),
#                 "DON_GIA": hhdvu.findtext('DGia'),
#                 "CK_LINE": hhdvu.findtext('STCKhau') if hhdvu is not None else "",
#                 "THANH_TIEN": hhdvu.findtext('ThTien'),
#                 "PT_VAT": hhdvu.findtext('TSuat') if hhdvu is not None else "",
#             })
#         return True, master_row, line_rows
#     except Exception as e:
#         return False, None, str(e)


# =========================
# WORKER THREAD: parse XML -> append vao all_master_rows + all_line_rows
# =========================
def converter_worker(xml_queue, ket_qua_convert, all_master_rows, all_line_rows, lock_rows, stop_signal, xml_map):
    """
    Worker thread chay song song voi tai XML.
    Parse XML theo xml_map, append vao 2 list dung chung.
    """
    while True:
        try:
            item = xml_queue.get(timeout=2)
        except queue.Empty:
            if stop_signal.is_set():
                break
            continue

        if item is None:
            xml_queue.task_done()
            break

        xml_path, ma_hd, MA_DONVI, mst_value, huong, thang, nam, khhd, shd, ten_base = item

        ok, master_row, line_rows = parse_xml_hoa_don(
            xml_path=xml_path,
            ma_hd=ma_hd,
            MA_DONVI=MA_DONVI,
            mst_value=mst_value,
            huong=huong,
            thang=thang,
            nam=nam,
            khhd=khhd,
            shd=shd,
            xml_map=xml_map
        )

        if ok:
            with lock_rows:
                all_master_rows.append(master_row)
                all_line_rows.extend(line_rows)
            ket_qua_convert["ok"] += 1
        else:
            ket_qua_convert["err"] += 1
            ket_qua_convert["loi"].append((ten_base, line_rows))  # line_rows la error message
            #print(f"      [Parse LOI] {ten_base}: {line_rows}")

        xml_queue.task_done()


def mo_trang_gdt(driver, run_log, url="https://hoadondientu.gdt.gov.vn", so_lan=5):
    for lan in range(1, so_lan + 1):
        try:
            driver.get(url)
        except WebDriverException as e:
            append_run_log(run_log, f"OPEN_SITE_GET_EXC lan={lan}: {e!r}")
            try: driver.get("about:blank")
            except Exception: pass
            time.sleep(min(3 * lan, 15))
            continue

        # Thanh cong = phan tu goc #__next cua trang GDT that su xuat hien
        try:
            WebDriverWait(driver, 12).until(
                lambda d: d.execute_script("return !!document.getElementById('__next');")
            )
            append_run_log(run_log, f"OPEN_SITE_OK lan={lan}")
            return True
        except TimeoutException:
            # Ghi lai trang dang hien ra (de biet la ERR gi)
            body_txt = ""
            try:
                body_txt = (driver.find_element(By.TAG_NAME, "body").text or "")[:150].replace("\n", " ")
            except Exception:
                pass
            append_run_log(run_log, f"OPEN_SITE_FAIL lan={lan} body={body_txt!r}")
            try: driver.get("about:blank")
            except Exception: pass
            time.sleep(min(3 * lan, 15))   # cho tang dan: 3,6,9,12,15s

    return False
# def converter_worker(xml_queue, ket_qua_convert, all_master_rows, all_line_rows, lock_rows, stop_signal):
#     """
#     Worker thread chay song song voi tai XML.
#     Parse XML, append rows vao 2 list dung chung.
#     Cuoi cung main thread se ghi 1 file Excel.
#     """
#     while True:
#         try:
#             item = xml_queue.get(timeout=2)
#         except queue.Empty:
#             if stop_signal.is_set():
#                 break
#             continue

#         if item is None:
#             xml_queue.task_done()
#             break

#         xml_path, ma_hd, MA_DONVI, mst_value, huong, thang, nam, khhd, shd, ten_base = item
#         ok, master_row, line_rows = parse_xml_hoa_don(xml_path, ma_hd, MA_DONVI, mst_value, huong, thang, nam, khhd, shd)
#         if ok:
#             master_row["HUONG"] = huong
#             for lr in line_rows:
#                 lr["HUONG"] = huong
#             with lock_rows:
#                 all_master_rows.append(master_row)
#                 all_line_rows.extend(line_rows)
#             ket_qua_convert["ok"] += 1
#         else:
#             ket_qua_convert["err"] += 1
#             ket_qua_convert["loi"].append((ten_base, line_rows))  # line_rows chua error msg
#             #print(f"      [Parse LOI] {ten_base}: {line_rows}")

#         xml_queue.task_done()


class tra_cuu_hdt:
    my_clsid = pythoncom.CreateGuid()
    _reg_clsid_ = my_clsid
    _reg_desc_ = "tra_cuu_hdt Server"
    _reg_progid_ = "tra_cuu_hdt.Server"
    _public_methods_ = ['xuat_hoa_don']

    # loai_xuat: "all" = tat ca, "ra" = chi ban ra, "vao" = chi mua vao

    # def xuat_hoa_don(self, mst_value, password_value, thang_bd, thang_kt, nam, save_dir, MA_DONVI, job_id,status, events, stagedir, loai_xuat="all"):

    def xuat_hoa_don(self, mst_value, password_value,tu_ngay,den_ngay,thang_bd, thang_kt, nam, save_dir, MA_DONVI, job_id, status, events, stagedir, loai_xuat="all", xml_map_path=None, tang_dan=False):
        paths = make_job_paths(save_dir, job_id, MA_DONVI)

        status_path = status
        events_path = events
        run_log = paths["run_log"]

        status = StatusWriter(status_path)
        events = EventWriter(events_path)

        append_run_log(run_log, f"JOB_ENTER xuat_hoa_don pid={os.getpid()} job_id={job_id} ma_donvi={MA_DONVI} mst={mst_value} loai_xuat={loai_xuat}")
        append_run_log(run_log, f"PATHS status={status_path} events={events_path} stage={stagedir}")

        if not xml_map_path:
            raise ValueError("Thieu xml_map_path")

        xml_map = load_xml_map_from_excel(xml_map_path)
        append_run_log(
            run_log,
            f"XML_MAP_LOADED path={xml_map_path} master={len(xml_map.get('master', []))} line={len(xml_map.get('line', []))}"
        )
        # paths = make_job_paths(save_dir, job_id,MA_DONVI)

        # # status = StatusWriter(paths["status"])
        # # events = EventWriter(paths["events"])
        # status = StatusWriter(status)
        # events = EventWriter(events)

        # NT-01: mọi "message" ghi ra status.json phải là tiếng Việt CÓ DẤU. Trước đây
        # viết không dấu nên giao diện web hiện "Dang dang nhap..." — người dùng tưởng
        # lỗi font, thực ra chuỗi nguồn vốn đã không dấu. StatusWriter ghi UTF-8 với
        # ensure_ascii=False sẵn rồi nên chỉ cần sửa chuỗi tại nguồn.
        #
        # loai_xuat / nam / thang_bd / thang_kt: web cần biết lượt này lấy hướng nào,
        # kỳ nào để hiện thành cột riêng thay vì bắt người đọc đoán từ câu message.
        state = {
            "job_id": job_id,
            "ma_donvi": MA_DONVI,
            "mst": mst_value,
            "loai_xuat": loai_xuat,
            "nam": nam,
            "thang_bd": thang_bd,
            "thang_kt": thang_kt,
            "state": "INIT",
            "message": "Khởi tạo",
            # Bo dem CONG DON CA JOB, dat trong state nen moi status.write ({**state,...})
            # deu mang theo. Truoc day chung chi nam trong dict cua tung lan write, nen
            # lan write cuoi cung khong co -> web giu lai so cu ("10/29") du da xong.
            "tong_hd": 0,          # tong hoa don phai tai, dem tu Excel danh sach
            "tai_ok": 0,           # tai ve duoc
            "khong_co_goc": 0,     # HTTP 500 "khong ton tai ho so goc" — HOP LE
            "loi_that": 0,         # 429 / 504 / mang hong — dang di tim
            # Bon so tren la TONG ca hai huong. Chay "ca vao ca ra" thi nhin so tong
            # khong biet file nao cua ben nao, HD khong co goc thuoc dau vao hay dau ra
            # (chot Truong 11/08). Tach rieng, van GIU khoa tong de ban C# cu doc duoc.
            "tong_hd_vao": 0, "tai_ok_vao": 0, "khong_co_goc_vao": 0, "loi_that_vao": 0,
            "tong_hd_ra": 0,  "tai_ok_ra": 0,  "khong_co_goc_ra": 0,  "loi_that_ra": 0,
            "nguon_ds": "",        # excel | search
            "stage_dbf_dir": stagedir,
            "run_log": run_log,
            "pid": os.getpid(),
            "alive": True,
            "started_at": datetime.datetime.now().isoformat(timespec="seconds")
        }
        # state = {
        #     "job_id": job_id,
        #     "ma_donvi": MA_DONVI,
        #     "mst": mst_value,
        #     "state": "INIT",
        #     "message": "Khởi tạo",
        #     "stage_dbf_dir": stagedir,
        #     "started_at": datetime.datetime.now().isoformat(timespec="seconds")
        # }

        # status.write(state)
        # events.log("LOGIN", ma_donvi=MA_DONVI)
        ##########
        status.write(state)
        events.log("LOGIN", ma_donvi=MA_DONVI)

        last_heartbeat = 0.0

        def set_state(new_state=None, message=None, **extra):
            if new_state is not None:
                state["state"] = new_state
            if message is not None:
                state["message"] = message

            payload = {
                **state,
                "alive": True,
                "pid": os.getpid(),
                **extra
            }
            status.write(payload)

        def heartbeat(message=None, force=False, **extra):
            nonlocal last_heartbeat
            now = time.time()
            if force or (now - last_heartbeat >= 10):
                payload = {
                    **state,
                    "alive": True,
                    "pid": os.getpid(),
                    "heartbeat_at": datetime.datetime.now().isoformat(timespec="seconds"),
                    **extra
                }
                if message:
                    payload["message"] = message
                status.write(payload)
                last_heartbeat = now
        ##########

        # temp_profile_dir = tempfile.mkdtemp(prefix="chrome_profile_")
        # chrome_options = webdriver.ChromeOptions()
        # chrome_options.add_argument("--headless=new")
        # chrome_options.add_argument("--window-size=1920,1080")
        # chrome_options.add_argument(f"--user-data-dir={temp_profile_dir}")
        # chrome_options.add_argument("--force-device-scale-factor=1")
        # chrome_options.add_argument("--high-dpi-support=1")
        # chrome_options.add_argument("--disable-gpu")
        # chrome_options.add_argument("--hide-scrollbars")
        # driver = webdriver.Chrome(options=chrome_options)

        append_run_log(run_log, "CHROME_PROFILE_CREATE")
        # Kill chromedriver/chrome con sot tu lan chay truoc (tranh khoa profile)
        os.system('taskkill /F /IM chromedriver.exe /T >NUL 2>&1')

        # Don sach cac profile rac "chrome_profile_*" con ton tai do lan chay truoc bi crash/kill
        try:
            temp_root = tempfile.gettempdir()
            for name in os.listdir(temp_root):
                if name.startswith("chrome_profile_"):
                    old_path = os.path.join(temp_root, name)
                    if os.path.isdir(old_path):
                        shutil.rmtree(old_path, ignore_errors=True)
        except Exception:
            pass

        # Uu tien cache local cua webdriver-manager, chi tai lai khi Chrome doi version
        os.environ.setdefault('WDM_LOCAL', '1')

        # Chrome + profile duoc tao MOI trong moi lan dang nhap (xem vong retry ben duoi)
        driver = None
        temp_profile_dir = None

        # ===== Khoi tao queue + worker thread =====
        xml_queue = queue.Queue()
        ket_qua_convert = {"ok": 0, "err": 0, "loi": []}
        all_master_rows = []   # Toan bo dong master (sheet ThongTinChung)
        all_line_rows = []     # Toan bo dong line (sheet HangHoa)
        lock_rows = threading.Lock()
        excel_summary_files = []   # Danh sach file Excel tong hop de bo sung NOXML va cap nhat TTHAI_HD        
        stop_signal = threading.Event()

        # worker = threading.Thread(
        #     target=converter_worker,
        #     args=(xml_queue, ket_qua_convert, all_master_rows, all_line_rows, lock_rows, stop_signal),
        #     daemon=True
        # )
        worker = threading.Thread(
            target=converter_worker,
            args=(xml_queue, ket_qua_convert, all_master_rows, all_line_rows, lock_rows, stop_signal, xml_map),
            daemon=True
        )
        worker.start()

        try:
            if not (1 <= thang_bd <= 12 and 1 <= thang_kt <= 12 and thang_bd <= thang_kt):
                status.write({**state, "state": "ERROR", "message": "Khoảng tháng không hợp lệ"})
                events.log("JOB_ERROR", error="invalid_month_range")
                return "invalid_month_range"
            khoang_cach = []
            if tu_ngay and den_ngay:
                tu_ngay = tu_ngay
                den_ngay = den_ngay
                khoang_cach.append((thang_bd, tu_ngay, den_ngay))
            else:
                for thang in range(thang_bd, thang_kt + 1):
                    tu_ngay = f"01/{thang:02d}/{nam}"
                    _, so_ngay = calendar.monthrange(nam, thang)
                    den_ngay = f"{so_ngay:02d}/{thang:02d}/{nam}"
                    
                    khoang_cach.append((thang, tu_ngay, den_ngay))
                
            # #print("đang mở trình duyệt và đăng nhập...")
            # events.log("LOGIN",message="đang mở trình duyệt và đăng nhập...")
            # t_start = time.time()
            # driver.get("https://hoadondientu.gdt.gov.vn/")

            #print("đang mở trình duyệt và đăng nhập...")
            # ====================================================================
            # DANG NHAP GDT VOI RETRY: that bai -> dong Chrome, mo lai, login lai
            # ====================================================================
            append_run_log(run_log, "LOGIN_OPEN_SITE_START")
            events.log("LOGIN", message="Dang mo trinh duyet va dang nhap...")

            class DangNhapSai(Exception):
                """Sai MST/mat khau, hoac captcha sai sap cham nguong khoa -> dung han, KHONG retry."""
                pass

            # GDT khoa tai khoan neu nhap sai captcha ~5 lan. Vong retry ngoai x so lan thu
            # captcha trong phien co the cong don -> dem TONG so lan captcha SAI da SUBMIT
            # tren toan bo qua trinh login va dung lai truoc khi cham nguong.
            GIOI_HAN_CAPTCHA_SAI = 4   # tong so lan submit captcha sai toi da (an toan duoi 5)
            wrong_captcha_total = 0

            def _tao_chrome():
                _profile = tempfile.mkdtemp(prefix="chrome_profile_")
                _opts = webdriver.ChromeOptions()
                _opts.add_argument("--headless=new")
                _opts.add_argument("--disable-http2")
                _opts.add_argument("--disable-quic")
                _opts.add_argument("--window-size=1920,1080")
                _opts.add_argument(f"--user-data-dir={_profile}")
                _opts.add_argument("--disable-gpu")
                _opts.add_argument("--no-sandbox")
                _opts.add_argument("--disable-dev-shm-usage")
                _opts.add_argument("--disable-blink-features=AutomationControlled")
                _opts.add_argument("--hide-scrollbars")
                _drv = webdriver.Chrome(
                    service=Service(ChromeDriverManager().install()),
                    options=_opts
                )
                return _drv, _profile

            def _login_mot_lan():
                """Mot lan dang nhap voi Chrome moi.
                Tra ve: token (str) | None (loi tam thoi -> retry) | raise DangNhapSai (vinh vien)."""
                nonlocal wrong_captcha_total
                drv = None
                profile = None
                try:
                    append_run_log(run_log, "CHROME_START")
                    drv, profile = _tao_chrome()
                    append_run_log(run_log, "CHROME_STARTED")

                    set_state("LOGIN", "Dang mo website GDT")
                    heartbeat("Dang mo website GDT", force=True)
                    if not mo_trang_gdt(drv, run_log):
                        return None
                    append_run_log(run_log, "LOGIN_OPEN_SITE_DONE")

                    wait = WebDriverWait(drv, 12)

                    time.sleep(2)
                    close_btns = drv.find_elements(By.CSS_SELECTOR, ".ant-modal-close, .ant-modal-close-x")
                    if close_btns:
                        close_btns[0].click()
                    else:
                        modal_wrap = drv.find_elements(By.CSS_SELECTOR, ".ant-modal-wrap")
                        if modal_wrap:
                            ActionChains(drv).move_to_element_with_offset(modal_wrap[0], 10, 10).click().perform()
                    time.sleep(1)

                    login_btn = wait.until(EC.element_to_be_clickable(
                        (By.XPATH, '//*[@id="__next"]/section/header/div[1]/div/div[7]')
                    ))
                    login_btn.click()
                    time.sleep(2)

                    mst_input = wait.until(EC.element_to_be_clickable(
                        (By.XPATH, '/html/body/div[2]/div/div[2]/div/div[2]/div[2]/form/div/div[1]/div/div[2]/div')
                    ))
                    password_input = wait.until(EC.element_to_be_clickable(
                        (By.XPATH, '/html/body/div[2]/div/div[2]/div/div[2]/div[2]/form/div/div[2]/div/div[2]/div')
                    ))
                    mst_input.click()
                    mst_input.find_element(By.TAG_NAME, "input").send_keys(mst_value)
                    password_input.click()
                    password_input.find_element(By.TAG_NAME, "input").send_keys(password_value)
                    events.log("LOGIN", message=f"Da nhap MST: {mst_value}")
                    status.write({**state, "state": "LOGIN", "message": "Đang đăng nhập cổng Tổng cục Thuế..."})

                    ocr = ddddocr.DdddOcr(show_ad=False)
                    CAPTCHA_IMG_XPATH = '/html/body/div[2]/div/div[2]/div/div[2]/div[2]/form/div/div[3]/div/div[2]/div/span/div/img'
                    CAPTCHA_INPUT_XPATH = '/html/body/div[2]/div/div[2]/div/div[2]/div[2]/form/div/div[4]/div/div[2]/div'
                    MAX_RETRIES = 4

                    for attempt in range(1, MAX_RETRIES + 1):
                        append_run_log(run_log, f"CAPTCHA_ATTEMPT_{attempt}_START")
                        time.sleep(1)

                        # Lay anh captcha - chiu duoc loi tam thoi (timeout/stale)
                        png = None
                        for _cap_try in range(3):
                            try:
                                captcha_img = WebDriverWait(drv, 5).until(
                                    EC.visibility_of_element_located((By.XPATH, CAPTCHA_IMG_XPATH))
                                )
                                png = captcha_img.screenshot_as_png
                                if png:
                                    break
                            except (StaleElementReferenceException, TimeoutException, WebDriverException) as e:
                                append_run_log(run_log, f"CAPTCHA_ATTEMPT_{attempt}_IMG_RETRY{_cap_try}: {type(e).__name__}")
                                time.sleep(1.5)
                        if png is None:
                            append_run_log(run_log, f"CAPTCHA_ATTEMPT_{attempt}_IMG_MISSING")
                            return None   # anh khong len -> server chet -> mo lai phien moi

                        im = Image.open(io.BytesIO(png)).convert('L')
                        im_con = ImageOps.autocontrast(im, cutoff=15)
                        im_con_big = im_con.resize((im.width * 3, im.height * 3), Image.LANCZOS)
                        buf = io.BytesIO()
                        im_con_big.save(buf, format='PNG')
                        captcha_text = ocr.classification(buf.getvalue()).upper()
                        captcha_text = re.sub(r'[^A-Z0-9]', '', captcha_text)   # bo ky tu rac (vd '一', '-')
                        append_run_log(run_log, f"CAPTCHA_ATTEMPT_{attempt}_OCR={captcha_text}")
                        # Captcha GDT = 6 ky tu chu/so. Sai do dai => chac chan OCR doc loi,
                        # doi anh captcha moi roi doc lai - KHONG bam dang nhap (khong tinh vao nguong khoa).
                        if len(captcha_text) != 6:
                            append_run_log(run_log, f"CAPTCHA_ATTEMPT_{attempt}_BAD_LEN len={len(captcha_text)}")
                            try:
                                drv.find_element(By.XPATH, CAPTCHA_IMG_XPATH).click()  # bam vao anh -> GDT doi captcha moi
                                time.sleep(0.8)
                            except Exception:
                                pass
                            continue
                        events.log("LOGIN", message=f"Lan thu {attempt}: Ma CAPCHA : {captcha_text} ")

                        # Nhap captcha + bam dang nhap - boc try de loi tam thoi khong vo phien
                        try:
                            cap_input = WebDriverWait(drv, 10).until(
                                EC.element_to_be_clickable((By.XPATH, CAPTCHA_INPUT_XPATH)))
                            cap_input.click()
                            input_el = cap_input.find_element(By.TAG_NAME, "input")
                            input_el.clear()
                            input_el.send_keys(captcha_text)
                            login_btn2 = drv.find_element(
                                By.XPATH, '/html/body/div[2]/div/div[2]/div/div[2]/div[2]/form/div/div[6]/button')
                            login_btn2.click()
                            time.sleep(1.5)
                        except (TimeoutException, StaleElementReferenceException, WebDriverException) as e:
                            append_run_log(run_log, f"CAPTCHA_ATTEMPT_{attempt}_INPUT_ERR: {type(e).__name__}")
                            return None   # transient -> mo lai phien moi

                        # Kiem tra ket qua dang nhap
                        try:
                            error_notis = drv.find_elements(By.CLASS_NAME, "ant-notification-notice-message")
                            if error_notis:
                                error_text = error_notis[0].text
                                if "Ma captcha khong dung" in error_text or "Mã captcha không đúng" in error_text:
                                    wrong_captcha_total += 1
                                    append_run_log(run_log, f"CAPTCHA_WRONG_TOTAL={wrong_captcha_total}/{GIOI_HAN_CAPTCHA_SAI}")
                                    if wrong_captcha_total >= GIOI_HAN_CAPTCHA_SAI:
                                        append_run_log(run_log, "CAPTCHA_NEAR_LOCK -> dung de tranh khoa tai khoan GDT")
                                        raise DangNhapSai("captcha_near_lock")
                                    try:
                                        drv.find_element(By.XPATH, CAPTCHA_IMG_XPATH).click()  # doi anh captcha moi
                                        time.sleep(0.8)
                                    except Exception:
                                        pass
                                    continue   # thu captcha moi (van trong phien hien tai)
                                elif "Ten dang nhap hoac mat khau khong dung" in error_text or "Tên đăng nhập hoặc mật khẩu không đúng" in error_text:
                                    append_run_log(run_log, f"CAPTCHA_ATTEMPT_{attempt}_SAI_MK")
                                    raise DangNhapSai("login_failed")
                            if len(drv.find_elements(By.ID, "username")) == 0:
                                status.write({**state, "state": "LOGIN_OK", "message": "Đăng nhập thành công"})
                                events.log("LOGIN_OK", message="DANG NHAP THANH CONG!")
                                append_run_log(run_log, f"LOGIN_OK_ON_ATTEMPT_{attempt}")
                                break   # thanh cong
                        except StaleElementReferenceException:
                            status.write({**state, "state": "LOGIN_OK", "message": "Đăng nhập thành công"})
                            events.log("LOGIN_OK", message="DANG NHAP THANH CONG!")
                            append_run_log(run_log, f"LOGIN_OK_ON_ATTEMPT_{attempt}")
                            break   # trang da chuyen -> thanh cong
                    else:
                        append_run_log(run_log, "CAPTCHA_FAILED_ALL")
                        return None   # het luot captcha -> phien moi se co captcha khac

                    # ----- Lay token jwt tu cookie -----
                    time.sleep(2)
                    append_run_log(run_log, "GET_TOKEN_FROM_COOKIE_START")
                    _token = ""
                    for cookie in drv.get_cookies():
                        if cookie['name'] == 'jwt':
                            _token = cookie['value']
                            break
                    append_run_log(run_log, f"GET_TOKEN_DONE has_token={bool(_token)}")
                    return _token or None

                finally:
                    # Dong Chrome + xoa profile DU thanh cong hay loi (tranh ro tien trinh/khoa profile)
                    if drv is not None:
                        try:
                            drv.quit()
                        except Exception:
                            pass
                    if profile:
                        shutil.rmtree(profile, ignore_errors=True)

            # ----- Vong retry ngoai: dong/mo lai Chrome, login lai toi N lan -----
            token = None
            _SO_LAN_LOGIN = 5   # so lan dong/mo lai Chrome khi server GDT chap chon (doi 4 neu muon)
            for _vong in range(1, _SO_LAN_LOGIN + 1):
                append_run_log(run_log, f"LOGIN_OUTER_TRY_{_vong}/{_SO_LAN_LOGIN}")
                try:
                    token = _login_mot_lan()
                    if token:
                        append_run_log(run_log, f"LOGIN_OUTER_OK vong={_vong}")
                        break
                    append_run_log(run_log, f"LOGIN_OUTER_EMPTY vong={_vong}")
                except DangNhapSai as _de:
                    _ly_do = str(_de)
                    if _ly_do == "captcha_near_lock":
                        append_run_log(run_log, "LOGIN_STOP_CAPTCHA_NEAR_LOCK")
                        status.write({**state, "state": "ERROR", "message": "Sai captcha nhiều lần — dừng để tránh bị khóa tài khoản"})
                        events.log("JOB_ERROR", error="captcha_near_lock")
                        return "captcha_failed"
                    append_run_log(run_log, "LOGIN_SAI_MK -> dung han, khong retry")
                    status.write({**state, "state": "ERROR", "message": "Đăng nhập thất bại"})
                    events.log("JOB_ERROR", error="login_failed")
                    return "login_failed"
                except Exception as _e:
                    append_run_log(run_log, f"LOGIN_OUTER_EXC vong={_vong}: {type(_e).__name__}: {_e}")
                if _vong < _SO_LAN_LOGIN:
                    _cho = min(10 * _vong, 40)   # backoff tang dan: 10s, 20s, 30s...
                    append_run_log(run_log, f"LOGIN_OUTER_BACKOFF {_cho}s")
                    time.sleep(_cho)

            if not token:
                append_run_log(run_log, "LOGIN_OUTER_GIVEUP")
                status.write({**state, "state": "ERROR", "message": "Đăng nhập thất bại sau nhiều lần thử"})
                events.log("JOB_ERROR", error="login_failed_after_retries")
                return "login_failed"

            # token co roi -> chay tiep phan export (giu nguyen, van trong khoi if nay)
            if token:


                ds_loai_hd = [
                    {
                        "huong": "RA", "hau_to": "",
                        "url": "https://hoadondientu.gdt.gov.vn/api/query/invoices/export-excel",
                        "ttxly": None, "type": None,
                        "search_url": "https://hoadondientu.gdt.gov.vn/api/query/invoices/sold",
                        "xml_url": "https://hoadondientu.gdt.gov.vn/api/query/invoices/export-xml",
                        "search_action": "T%C3%ACm%20ki%E1%BA%BFm%20(h%C3%B3a%20%C4%91%C6%A1n%20b%C3%A1n%20ra)",
                        "xml_action": "Xu%E1%BA%A5t%20xml%20(h%C3%B3a%20%C4%91%C6%A1n%20b%C3%A1n%20ra)",
                        "action": "Xu%E1%BA%A5t%20h%C3%B3a%20%C4%91%C6%A1n%20(h%C3%B3a%20%C4%91%C6%A1n%20b%C3%A1n%20ra)",
                    },
                    {
                        "huong": "RA", "hau_to": "_MTT",
                        "url": "https://hoadondientu.gdt.gov.vn/api/sco-query/invoices/export-excel",
                        "ttxly": None, "type": None,
                        "search_url": "https://hoadondientu.gdt.gov.vn/api/sco-query/invoices/sold",
                        "xml_url": "https://hoadondientu.gdt.gov.vn/api/sco-query/invoices/export-xml",
                        "search_action": "T%C3%ACm%20ki%E1%BA%BFm%20(h%C3%B3a%20%C4%91%C6%A1n%20m%C3%A1y%20t%C3%ADnh%20ti%E1%BB%81n%20b%C3%A1n%20ra)",
                        "xml_action": "Xu%E1%BA%A5t%20xml%20(h%C3%B3a%20%C4%91%C6%A1n%20m%C3%A1y%20t%C3%ADnh%20ti%E1%BB%81n%20b%C3%A1n%20ra)",
                        "action": "Xu%E1%BA%A5t%20h%C3%B3a%20%C4%91%C6%A1n%20(h%C3%B3a%20%C4%91%C6%A1n%20b%C3%A1n%20ra)",
                    },
                    {
                        "huong": "VAO", "hau_to": "_CM",
                        "url": "https://hoadondientu.gdt.gov.vn/api/query/invoices/export-excel-sold",
                        "ttxly": "5", "type": "purchase",
                        "search_url": "https://hoadondientu.gdt.gov.vn/api/query/invoices/purchase",
                        "xml_url": "https://hoadondientu.gdt.gov.vn/api/query/invoices/export-xml",
                        "search_action": "T%C3%ACm%20ki%E1%BA%BFm%20(h%C3%B3a%20%C4%91%C6%A1n%20mua%20v%C3%A0o)",
                        "xml_action": "Xu%E1%BA%A5t%20xml%20(h%C3%B3a%20%C4%91%C6%A1n%20mua%20v%C3%A0o)",
                        "action": "Xu%E1%BA%A5t%20h%C3%B3a%20%C4%91%C6%A1n%20(h%C3%B3a%20%C4%91%C6%A1n%20mua%20v%C3%A0o)",
                    },
                    {
                        "huong": "VAO", "hau_to": "_KM",
                        "url": "https://hoadondientu.gdt.gov.vn/api/query/invoices/export-excel-sold",
                        "ttxly": "6", "type": "purchase",
                        "search_url": "https://hoadondientu.gdt.gov.vn/api/query/invoices/purchase",
                        "xml_url": "https://hoadondientu.gdt.gov.vn/api/query/invoices/export-xml",
                        "search_action": "T%C3%ACm%20ki%E1%BA%BFm%20(h%C3%B3a%20%C4%91%C6%A1n%20mua%20v%C3%A0o)",
                        "xml_action": "Xu%E1%BA%A5t%20xml%20(h%C3%B3a%20%C4%91%C6%A1n%20mua%20v%C3%A0o)",
                        "action": "Xu%E1%BA%A5t%20h%C3%B3a%20%C4%91%C6%A1n%20(h%C3%B3a%20%C4%91%C6%A1n%20mua%20v%C3%A0o)",
                    },
                    {
                        "huong": "VAO", "hau_to": "_MTT",
                        "url": "https://hoadondientu.gdt.gov.vn/api/sco-query/invoices/export-excel-sold",
                        "ttxly": "8", "type": "purchase",
                        "search_url": "https://hoadondientu.gdt.gov.vn/api/sco-query/invoices/purchase",
                        "xml_url": "https://hoadondientu.gdt.gov.vn/api/sco-query/invoices/export-xml",
                        "search_action": "T%C3%ACm%20ki%E1%BA%BFm%20(h%C3%B3a%20%C4%91%C6%A1n%20m%C3%A1y%20t%C3%ADnh%20ti%E1%BB%81n%20mua%20v%C3%A0o)",
                        "xml_action": "Xu%E1%BA%A5t%20xml%20(h%C3%B3a%20%C4%91%C6%A1n%20m%C3%A1y%20t%C3%ADnh%20ti%E1%BB%81n%20mua%20v%C3%A0o)",
                        "action": "Xu%E1%BA%A5t%20h%C3%B3a%20%C4%91%C6%A1n%20(h%C3%B3a%20%C4%91%C6%A1n%20m%C3%A1y%20t%C3%ADnh%20ti%E1%BB%81n%20mua%20v%C3%A0o)",
                    },
                ]

                if loai_xuat == "ra":
                    ds_loai_hd = [l for l in ds_loai_hd if l["huong"] == "RA"]
                elif loai_xuat == "vao":
                    ds_loai_hd = [l for l in ds_loai_hd if l["huong"] == "VAO"]

                if not os.path.exists(save_dir):
                    os.makedirs(save_dir, exist_ok=True)

                status.write({**state, "state": "DOWNLOAD", "message": "Bắt đầu tải hóa đơn"})
                events.log("DOWNLOAD_START", loai_xuat=loai_xuat)

                def make_headers(action):
                    return {
                        "Authorization": f"Bearer {token}",
                        "Accept": "application/json, text/plain, */*",
                        "Accept-Language": "vi",
                        "Origin": "https://hoadondientu.gdt.gov.vn",
                        "Referer": "https://hoadondientu.gdt.gov.vn/",
                        "End-Point": "/tra-cuu/tra-cuu-hoa-don",
                        "Action": action,
                    }

                def tai_hd(hd, xml_url, xml_action, huong, thang, sub_dir,  MA_DONVI):
                    """Tai XML+HTML, luu vao sub_dir. Tra ve duong dan XML neu thanh cong."""
                    # Them MST nguoi ban (nbmst) vao ten file de tranh GHI DE:
                    # (khhdon, shdon) chi duy nhat THEO TUNG nguoi ban, nen 2 nha cung
                    # cap khac nhau van co the trung khhdon+shdon -> trung ten file.
                    # Dat nbmst TRUOC khhdon_shdon de logic retry (rsplit("_", 2)) van
                    # lay dung (khhdon, shdon) o 2 phan cuoi.
                    nbmst_an_toan = re.sub(r'[\\/:*?"<>|]', '', str(hd.get('nbmst', '')).strip())
                    ten_base = f"{MA_DONVI}_{huong}_T{thang}_{nbmst_an_toan}_{hd['khhdon']}_{hd['shdon']}"
                    url = f"{xml_url}?nbmst={hd['nbmst']}&khhdon={hd['khhdon']}&shdon={hd['shdon']}&khmshdon={hd['khmshdon']}"
                    headers = make_headers(xml_action)
                    ds_loi_chitiet = []
                    xml_path = None
                    for lan in range(1, 6):
                        try:
                            resp = requests.get(url, headers=headers, timeout=90)
                            if resp.status_code == 200:
                                try:
                                    with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
                                        for name in zf.namelist():
                                            ext = os.path.splitext(name)[1].lower()
                                            if ext in ('.xml', '.html', '.htm'):
                                                out_path = os.path.join(sub_dir, f"{ten_base}{ext}")
                                                with open(out_path, "wb") as f:
                                                    f.write(zf.read(name))
                                                if ext == '.xml':
                                                    xml_path = out_path
                                except zipfile.BadZipFile:
                                    out_path = os.path.join(sub_dir, f"{ten_base}.xml")
                                    with open(out_path, "wb") as f:
                                        f.write(resp.content)
                                    xml_path = out_path
                                return {"ok": True, "ten_base": ten_base, "xml_path": xml_path}
                            elif resp.status_code == 429:
                                ds_loi_chitiet.append(f"Lan {lan}: HTTP 429")
                                time.sleep(5 * lan)
                                #print(f"      {ten_base}: bi 429, doi {5 * lan}s...")
                                events.log("ERROR",message=f"      {ten_base}: bi 429, doi {5 * lan}s...")
                            elif resp.status_code == 500:
                                try:
                                    err_msg = resp.json().get("message", resp.text[:200])
                                except:
                                    err_msg = resp.text[:200] if resp.text else "Khong co thong tin"
                                if "Không tồn tại hồ sơ gốc" in str(err_msg):
                                    return {"ok": False, "ma": ten_base, "lydo": f"HTTP 500 - {err_msg}"}
                                ds_loi_chitiet.append(f"Lan {lan}: HTTP 500 - {err_msg}")
                                time.sleep(3)
                            elif resp.status_code == 504:
                                try:
                                    err_msg = resp.json().get("message", resp.text[:200])
                                except:
                                    err_msg = resp.text[:200] if resp.text else "Khong co thong tin"
                                ds_loi_chitiet.append(f"Lan {lan}: HTTP 504 - {err_msg}")
                                time.sleep(3)
                            else:
                                return {"ok": False, "ma": ten_base, "lydo": f"HTTP {resp.status_code}"}
                        except Exception as e:
                            ds_loi_chitiet.append(f"Lan {lan}: Exception - {e}")
                            time.sleep(3)
                    return {"ok": False, "ma": ten_base, "lydo": "That bai sau 5 lan thu: " + " | ".join(ds_loi_chitiet)}

                # result = []
                # for thang, tu_ngay, den_ngay in khoang_cach:
                #   for loai in ds_loai_hd:
                result = []
                # Cờ dừng êm. Bật rồi thì thoát khỏi mọi vòng lặp, NHƯNG vẫn đi qua
                # khối finally của lượt đang dở để chốt cột "Đường dẫn XML".
                dung_theo_yeu_cau = False

                for thang, tu_ngay, den_ngay in khoang_cach:
                  job_id = f"T{thang}_{nam}_{MA_DONVI}"
                  paths = make_job_paths(save_dir, job_id,MA_DONVI)
                  if nen_dung(paths["job_dir"]):
                      dung_theo_yeu_cau = True
                      append_run_log(run_log, f"DUNG_EM truoc thang={thang}")
                      break
                  heartbeat(f"Đang sử lý thang {thang}", force=True, current_month=thang)
                  append_run_log(run_log, f"MONTH_START thang={thang} tu_ngay={tu_ngay} den_ngay={den_ngay}")

                  for loai in ds_loai_hd:
                    if dung_theo_yeu_cau or nen_dung(paths["job_dir"]):
                        dung_theo_yeu_cau = True
                        append_run_log(run_log,
                            f"DUNG_EM truoc {loai['huong']}{loai['hau_to']} T{thang}")
                        break

                    # Khởi tạo TRƯỚC try: khối finally ở cuối phải chạy được kể cả khi
                    # lỗi nổ ngay dòng đầu, không thì nó ném NameError và che mất lỗi thật.
                    sub_dir = None
                    ten_file = None
                    xml_paths_cho_excel = {}
                    ket_qua_tai = {}

                    heartbeat(
                        f"Đang xử lý {loai['huong']}{loai['hau_to']} T{thang}",
                        force=True,
                        current_month=thang,
                        current_huong=loai["huong"],
                        current_suffix=loai["hau_to"]
                    )
                    append_run_log(run_log, f"TYPE_START thang={thang} huong={loai['huong']} hau_to={loai['hau_to']}")
                    try:
                        search_param = f"tdlap=ge={tu_ngay}T00:00:00;tdlap=le={den_ngay}T23:59:59"
                        if loai["ttxly"]:
                            search_param += f";ttxly=={loai['ttxly']}    "
                        api_url = f"{loai['url']}?sort=tdlap:desc&search={search_param}"
                        if loai["type"]:
                            api_url += f"&type={loai['type']}"

                        headers = {
                            "Authorization": f"Bearer {token}",
                            "Accept": "application/json, text/plain, */*",
                            "Accept-Language": "vi",
                            "Origin": "https://hoadondientu.gdt.gov.vn",
                            "Referer": "https://hoadondientu.gdt.gov.vn/",
                            "End-Point": "/tra-cuu/tra-cuu-hoa-don",
                            "Action": loai["action"],
                        }

                        # Duong dan: raw_dir\MST_MA_DONVI\(MV,BR)\NAM{nam}\HD_xx_Tx
                        huong_dir = "RA" if loai["huong"] == "RA" else "VAO"
                        # sub_dir = os.path.join(paths["raw_dir"], f"{MA_DONVI}_{mst_value}", huong_dir, f"NAM{nam}", f"HD_{huong_dir}_T{thang}")
                        sub_dir = os.path.join(paths["raw_dir"], huong_dir)
                        if not os.path.exists(sub_dir):
                            os.makedirs(sub_dir, exist_ok=True)

                        # Tai file Excel danh sach
                        ten_file = f"HD_{loai['huong']}_{MA_DONVI}_T{thang}{loai['hau_to']}.xlsx"
                        #print(f"Dang tai: HD_{loai['huong']}{loai['hau_to']} ({tu_ngay} -> {den_ngay})")
                        status.write({**state, "state": "EXCEL", "message": f"Tải Excel: HD_{loai['huong']}{loai['hau_to']} T{thang}"})
                        events.log("EXCEL_FETCH", huong=loai['huong'], hau_to=loai['hau_to'], thang=thang)
                        resp = None
                        append_run_log(run_log, f"EXCEL_FETCH_START url={api_url}") ###

                        for lan_thu in range(1, 4):
                            try:                           
                                resp = requests.get(api_url, headers=headers, timeout=90)
                                append_run_log(run_log, f"EXCEL_FETCH_RESPONSE status_code={resp.status_code if resp is not None else 'None'}")
                                if resp.status_code == 200:
                                    break
                            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e_net:
                                #print(f"   [Excel] Lan {lan_thu} timeout/connection: {e_net}")
                                events.log("EXCELL",message=f"   [Excel] Lan {lan_thu} timeout/connection: {e_net}")
                                resp = None
                            time.sleep(3)
                        if resp is not None and resp.status_code == 200:
                            save_path = os.path.join(sub_dir, ten_file)
                            # Che do tang dan: hop nhat ban vua tai voi ban dang co, giu
                            # lai hai cot chu thich. Hop nhat that bai (khong co file cu,
                            # trung khoa, cong doi cau truc...) thi ghi thang nhu cu -
                            # cham hon nhung khong bao gio bo sot hoa don.
                            tk_hop_nhat = None
                            if tang_dan and os.path.exists(save_path):
                                tk_hop_nhat = hop_nhat_excel(
                                    save_path, resp.content, thang, paths["job_dir"], run_log)
                            if tk_hop_nhat is None:
                                with open(save_path, "wb") as f:
                                    f.write(resp.content)
                            else:
                                append_run_log(run_log,
                                    f"HOP_NHAT {ten_file} tong={tk_hop_nhat['tong']} "
                                    f"giu={tk_hop_nhat['giu']} "
                                    f"doi_ban_chat={tk_hop_nhat['doi_ban_chat']} "
                                    f"mat={tk_hop_nhat['mat']}")
                            #print(f"Da luu: {save_path}")
                            events.log("EXCEL",message=f"Da luu: {save_path}")
                            append_run_log(run_log, f"EXCEL_SAVED {save_path}")
                            result.append(save_path)
                            excel_summary_files.append({
                                "path": save_path,
                                "huong": loai["huong"]
                            })
                        # === Search danh sach hoa don ===
                        headers = make_headers(loai["search_action"])
                        #print(f"\nTim kiem: HD_{loai['huong']}{loai['hau_to']} T{thang}...")
                        ds_hd = []
                        seen_ids = set()
                        page = 0
                        page_state = None
                        da_co_tu_excel = False
                        nguon_ds = "search"

                        # Uu tien dung chinh Excel vua tai: no da co du khmshdon/khhdon/
                        # shdon/nbmst de dung URL tai XML. Excel hong hoac doi khuon thi
                        # quay ve search chu khong lam chet ca thang.
                        excel_ds_path = os.path.join(sub_dir, ten_file)
                        if DUNG_EXCEL_THAY_SEARCH and os.path.exists(excel_ds_path):
                            try:
                                ds_hd = build_ds_hd_tu_excel(excel_ds_path)
                                da_co_tu_excel = True
                                nguon_ds = "excel"
                                append_run_log(run_log, f"DS_HD_FROM_EXCEL file={excel_ds_path} count={len(ds_hd)}")
                                status.write({**state, "state": "EXCEL",
                                              "message": f"Doc danh sach tu Excel: {len(ds_hd)} hoa don"})
                            except Exception as e_ex:
                                append_run_log(run_log, f"DS_HD_FROM_EXCEL_ERROR {short_exc(e_ex)} -> fallback search")
                                ds_hd = []
                                da_co_tu_excel = False
                        state["nguon_ds"] = nguon_ds

                        # Bo nhung hoa don DA co XML hoac DA xac nhan khong co goc.
                        # Chi loc khi danh sach den TU EXCEL: neu phai quay ve search thi
                        # file Excel khong dang tin, loc theo no la bo sot.
                        if tang_dan and da_co_tu_excel:
                            ds_hd, dung_lai, so_bo, so_khong_goc = loc_hoa_don_can_tai(
                                ds_hd, excel_ds_path, run_log)

                            # Hoa don bo tai nhung file XML van con tren dia: van PHAI
                            # day qua hang doi chuyen doi, khong thi no vang mat khoi
                            # Excel tong va buoc nap khong nhin thay -> file nam ly lai
                            # trong raw\. Cai ta tiet kiem la LUOT GOI MANG, khong phai
                            # cong doan doc file.
                            for hd_cu, dd_cu in dung_lai:
                                mst_ph = str(hd_cu.get('nbmst') or mst_value).strip()
                                xml_queue.put((
                                    dd_cu,
                                    f"{loai['huong']}_{mst_ph}_{hd_cu['khhdon']}_{hd_cu['shdon']}",
                                    MA_DONVI, mst_value, loai["huong"], thang, nam,
                                    hd_cu['khhdon'], hd_cu['shdon'],
                                    os.path.splitext(os.path.basename(dd_cu))[0]
                                ))

                            append_run_log(run_log,
                                f"TANG_DAN bo_qua={so_bo} (khong_co_goc={so_khong_goc}, "
                                f"dung_lai_file_cu={len(dung_lai)}) con_tai={len(ds_hd)}")
                            status.write({**state, "state": "EXCEL",
                                          "message": f"Tăng dần: dùng lại {len(dung_lai)} file có sẵn, "
                                                     f"tải mới {len(ds_hd)}"})

                        while not da_co_tu_excel:
                            search_url = f"{loai['search_url']}?sort=tdlap:desc&size=50&search={search_param}"
                            heartbeat(
                                f"Search danh sach {loai['huong']}{loai['hau_to']} T{thang} trang {page}",
                                current_month=thang,
                                current_page=page
                            )
                            append_run_log(run_log, f"SEARCH_PAGE_START page={page} url={search_url if 'search_url' in locals() else ''}")
                            if page_state:
                                search_url += f"&state={page_state}"
                            resp = None
                            for lan in range(1, 5):
                                try:
                                    resp = requests.get(search_url, headers=headers, timeout=90)
                                    append_run_log(run_log, f"SEARCH_PAGE_RESPONSE page={page} status_code={resp.status_code if resp is not None else 'None'}")
                                    if resp.status_code == 200:
                                        break
                                except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e_net:
                                    #print(f"   [Search] Trang {page} lan {lan} timeout/connection: {e_net}")
                                    resp = None
                                time.sleep(3 * lan)
                            if resp is None or resp.status_code != 200:
                                #print(f"   Loi search trang {page}: {resp.status_code if resp else 'timeout'}")
                                break
                            data = resp.json()
                            datas = data.get("datas", [])
                            total = data.get("total", 0)
                            page_state = data.get("state", None)
                            #if page == 0:
                                #print(f"   Tong so hoa don: {total}")
                            if not datas:
                                break
                            for hd in datas:
                                hd_key = f"{hd.get('khhdon')}_{hd.get('shdon')}"
                                if hd_key not in seen_ids:
                                    seen_ids.add(hd_key)
                                    ds_hd.append(hd)
                            #print(f"   Trang {page}: tong: {len(ds_hd)}/{total}")
                            if total and len(ds_hd) >= total:
                                break
                            if len(datas) < 50:
                                break
                            page += 1
                            time.sleep(1)

                        if not ds_hd:
                            #print(f"   Khong co hoa don nao")
                            continue

                        #print(f"   Tong: {len(ds_hd)} hoa don. Tai XML + HTML (convert song song)...")
                        # _h = "vao" | "ra": hau to cua bo dem tach theo huong. Lay tu
                        # loai["huong"] chu khong tu loai_xuat — chay "all" thi loai_xuat
                        # la "all" cho ca hai luot, khong phan biet duoc.
                        _h = loai["huong"].lower()
                        state["tong_hd"] = state.get("tong_hd", 0) + len(ds_hd)
                        state[f"tong_hd_{_h}"] = state.get(f"tong_hd_{_h}", 0) + len(ds_hd)
                        status.write({**state, "state": "XML",
                                      "message": f"Tải XML: HD_{loai['huong']}{loai['hau_to']} T{thang}",
                                      "total": len(ds_hd), "downloaded": 0})
                        events.log("XML_FETCH_START", huong=loai['huong'], hau_to=loai['hau_to'], thang=thang, total=len(ds_hd))

                        # Tai tuan tu, push XML vao queue de worker convert song song
                        thanh_cong = 0
                        ds_loi = []
                        xml_paths_cho_excel = {}  # { (khhd, shd): xml_path }
                        # { (khhd, shd): "KHÔNG CÓ GỐC" } — ket qua CUOI CUNG, khong phai
                        # that bai tam thoi. Khong ghi lai thi 15-22% hoa don (nhom khong
                        # ma: dien, nuoc, vien thong) bi thu lai moi ngay, mai mai.
                        ket_qua_tai = {}
                        for i, hd in enumerate(ds_hd, 1):
                            # Dừng êm: chỉ thoát VÒNG TẢI, không nhảy thẳng ra ngoài —
                            # phải đi qua finally để ghi cột "Đường dẫn XML" cho những
                            # hóa đơn vừa tải xong, không thì lượt sau tải lại từ đầu.
                            if nen_dung(paths["job_dir"]):
                                dung_theo_yeu_cau = True
                                append_run_log(run_log,
                                    f"DUNG_EM giua vong tai {loai['huong']}{loai['hau_to']} "
                                    f"T{thang} tai {i}/{len(ds_hd)}")
                                status.write({**state, "state": "DA_DUNG",
                                              "message": f"Đang dừng — đã tải {i-1}/{len(ds_hd)}, "
                                                         f"chốt đường dẫn XML"})
                                break
                            heartbeat(
                                f"Tai XML {loai['huong']}{loai['hau_to']} T{thang}: {i}/{len(ds_hd)}",
                                total=len(ds_hd),
                                downloaded=i - 1,
                                convert_ok=ket_qua_convert["ok"],
                                convert_err=ket_qua_convert["err"],
                                queue_size=xml_queue.qsize(),
                                current_month=thang,
                                current_huong=loai["huong"]
                            )
                        # for i, hd in enumerate(ds_hd, 1):

                            kq = tai_hd(hd, loai["xml_url"], loai["xml_action"], loai["huong"], thang, sub_dir,  MA_DONVI)
                            if kq["ok"]:
                                append_run_log(run_log, f"XML_OK {kq.get('ten_base','')}")
                            else:
                                append_run_log(run_log, f"XML_FAIL {kq.get('ma','')} reason={kq.get('lydo','')}")

                            time.sleep(1)
                            if kq["ok"]:
                                thanh_cong += 1
                                if kq.get("xml_path"):
                                    xml_paths_cho_excel[(str(hd['khhdon']), str(hd['shdon']))] = kq["xml_path"]
                                    # ma_hd = f"{MA_DONVI}_{mst_value}_{hd['khhdon']}_{hd['shdon']}"
                                    # MỚI:
                                    mst_ph = str(hd.get('nbmst') or mst_value).strip()
                                    # ma_hd = f"{huong}_{mst_ph}_{hd['khhdon']}_{hd['shdon']}"
                                    ma_hd = f"{loai['huong']}_{mst_ph}_{hd['khhdon']}_{hd['shdon']}"
                                    xml_queue.put((
                                        kq["xml_path"], ma_hd, MA_DONVI, mst_value,
                                        loai["huong"], thang, nam, hd['khhdon'], hd['shdon'],
                                        kq["ten_base"]
                                    ))
                            else:
                                ds_loi.append(kq)
                                # HTTP 500 "khong ton tai ho so goc" la ket qua CUOI CUNG,
                                # khong phai loi tam thoi: cong khong giu ban goc cho hoa
                                # don khong ma. Danh dau de khoi thu lai moi ngay.
                                if la_khong_co_goc(kq.get("lydo", "")):
                                    ket_qua_tai[(str(hd['khhdon']).strip(),
                                                 str(hd['shdon']).strip())] = KQ_KHONG_GOC
                            if i % 10 == 0 or i == len(ds_hd):
                                #print(f"   Da tai {i}/{len(ds_hd)} (OK: {thanh_cong}, Loi: {len(ds_loi)}, Queue convert: {xml_queue.qsize()})")
                                # Tach 500-khong-co-goc khoi loi that ngay tu day, de nguoi
                                # doc biet cai nao dang di tim, cai nao chi la hoa don dien
                                # nuoc vien thong von khong co ban goc tren cong.
                                so_khong_goc = sum(1 for l in ds_loi if la_khong_co_goc(l.get("lydo", "")))
                                status.write({**state, "state": "XML",
                                              "message": f"HD_{loai['huong']}{loai['hau_to']} T{thang}: {i}/{len(ds_hd)}",
                                              "total": len(ds_hd), "downloaded": i,
                                              "ok": thanh_cong, "err": len(ds_loi),
                                              "tai_ok": state.get("tai_ok", 0) + thanh_cong,
                                              "khong_co_goc": state.get("khong_co_goc", 0) + so_khong_goc,
                                              "loi_that": state.get("loi_that", 0) + len(ds_loi) - so_khong_goc,
                                              "convert_ok": ket_qua_convert["ok"], "convert_err": ket_qua_convert["err"]})
                                append_run_log(
                                    run_log,
                                    f"XML_PROGRESS {i}/{len(ds_hd)} ok={thanh_cong} err={len(ds_loi)} queue={xml_queue.qsize()} convert_ok={ket_qua_convert['ok']} convert_err={ket_qua_convert['err']}"
                                )

                        # Retry loi
                        if ds_loi:
                            ds_retry = [l for l in ds_loi if "hồ sơ gốc" not in str(l.get("lydo", ""))]
                            if ds_retry:
                                #print(f"   Thu lai {len(ds_retry)} hoa don loi (doi 30s)...")
                                append_run_log(run_log,f"   Thu lai {len(ds_retry)} hoa don loi (doi 30s)...")
                                time.sleep(30)

                                # FIX: Lookup chinh xac theo (khhdon, shdon) - tranh substring match nham
                                hd_index = {(str(h['khhdon']), str(h['shdon'])): h for h in ds_hd}

                                ds_loi_final = []
                                for l in ds_retry:
                                    # FIX: Tach khhdon, shdon tu ten_base thay vi substring search
                                    parts = l["ma"].rsplit("_", 2)
                                    key = (parts[-2], parts[-1]) if len(parts) >= 2 else None
                                    hd_retry = hd_index.get(key) if key else None

                                    if hd_retry:
                                        kq = tai_hd(hd_retry, loai["xml_url"], loai["xml_action"], loai["huong"], thang, sub_dir, MA_DONVI)
                                        time.sleep(2)
                                        if kq["ok"]:
                                            thanh_cong += 1
                                            if kq.get("xml_path"):
                                                # xml_paths_cho_excel[(str(hd_retry['khhdon']), str(hd_retry['shdon']))] = kq["xml_path"]
                                                # ma_hd = f"{MA_DONVI}_{mst_value}_{hd_retry['khhdon']}_{hd_retry['shdon']}"
                                                mst_ph = str(hd_retry.get('nbmst') or mst_value).strip()
                                                ma_hd = f"{loai['huong']}_{mst_ph}_{hd_retry['khhdon']}_{hd_retry['shdon']}"
                                                xml_queue.put((
                                                    kq["xml_path"], ma_hd, MA_DONVI, mst_value,
                                                    loai["huong"], thang, nam, hd_retry['khhdon'], hd_retry['shdon'],
                                                    kq["ten_base"]
                                                ))
                                            #print(f"      Retry OK: {kq['ten_base']}")  # FIX: in ten thuc su retry
                                            append_run_log(run_log, f"      Retry OK: {kq['ten_base']}")
                                        else:
                                            ds_loi_final.append(kq)
                                    else:
                                        ds_loi_final.append(l)
                                ds_loi_500 = [l for l in ds_loi if "hồ sơ gốc" in str(l.get("lydo", ""))]
                                ds_loi = ds_loi_final + ds_loi_500


                        # Chot so lieu cua luot nay vao bo dem ca job. Lam sau retry nen
                        # con so la ket qua CUOI CUNG, khong phai anh chup giua chung.
                        so_khong_goc = sum(1 for l in ds_loi if la_khong_co_goc(l.get("lydo", "")))
                        so_loi_that = len(ds_loi) - so_khong_goc
                        state["tai_ok"] = state.get("tai_ok", 0) + thanh_cong
                        state["khong_co_goc"] = state.get("khong_co_goc", 0) + so_khong_goc
                        state["loi_that"] = state.get("loi_that", 0) + so_loi_that
                        # Cùng số liệu, tách theo hướng — xem giải thích ở khối khởi tạo state
                        state[f"tai_ok_{_h}"] = state.get(f"tai_ok_{_h}", 0) + thanh_cong
                        state[f"khong_co_goc_{_h}"] = state.get(f"khong_co_goc_{_h}", 0) + so_khong_goc
                        state[f"loi_that_{_h}"] = state.get(f"loi_that_{_h}", 0) + so_loi_that
                        append_run_log(run_log,
                            f"LUOT_DONE {loai['huong']}{loai['hau_to']} T{thang} "
                            f"tong={len(ds_hd)} ok={thanh_cong} khong_goc={so_khong_goc} "
                            f"loi_that={len(ds_loi) - so_khong_goc} nguon={state.get('nguon_ds')}")

                        if ds_loi:
                            log_path = os.path.join(paths["job_dir"], f"LOI_TAI_{loai['huong']}{loai['hau_to']}_T{thang}.txt")
                            with open(log_path, "w", encoding="utf-8") as f:
                                for loi in ds_loi:
                                    f.write(f"{loi['ma']} - {loi['lydo']}\n")
                            #print(f"   {len(ds_loi)} hoa don loi, xem: {log_path}")
                            events.log("DOWNLOAD_ERRORS", count=len(ds_loi), huong=loai['huong'], thang=thang)

                        #print(f"   Hoan tat tai: {thanh_cong}/{len(ds_hd)} thanh cong")

                        result.append(sub_dir)
                    except Exception as e_loai:
                        #print(f"   [BO QUA] Loi xu ly HD_{loai['huong']}{loai['hau_to']} T{thang}: {e_loai}")
                        logging.error(f"Loi HD_{loai['huong']}{loai['hau_to']} T{thang}: {e_loai}", exc_info=True)
                        events.log("LOAI_ERROR", huong=loai['huong'], hau_to=loai['hau_to'], thang=thang, error=str(e_loai))
                        status.write({**state, "state": "LOAI_ERROR", "message": f"Bỏ qua HD_{loai['huong']}{loai['hau_to']} T{thang}: {e_loai}"})
                        continue
                    finally:
                        # Chốt cột "Đường dẫn XML" DÙ CÓ LỖI hay bị dừng giữa chừng.
                        # Trước đây khối này nằm trong try, nên nhánh `except ... continue`
                        # nhảy qua nó: một lỗi mạng ở hóa đơn cuối là mất dấu toàn bộ XML
                        # của cả lượt, lượt sau tải lại từ đầu.
                        # Trống cả hai bản đồ thì đừng gọi — mở/lưu workbook mất ~2 giây
                        # mà không ghi được gì.
                        if sub_dir and ten_file and (xml_paths_cho_excel or ket_qua_tai):
                            excel_server_path = os.path.join(sub_dir, ten_file)
                            if os.path.exists(excel_server_path):
                                ok_ghi = them_cot_xml_path(
                                    excel_server_path, xml_paths_cho_excel,
                                    ket_qua_tai=ket_qua_tai)
                                append_run_log(run_log,
                                    f"GHI_DUONG_DAN {'ok' if ok_ghi else 'LOI'} "
                                    f"duong_dan={len(xml_paths_cho_excel)} "
                                    f"ket_qua={len(ket_qua_tai)} "
                                    f"file={os.path.basename(excel_server_path)}")

                    # Dừng êm: đường dẫn đã chốt ở finally trên, giờ mới thoát vòng loại
                    if dung_theo_yeu_cau:
                        break

                  if dung_theo_yeu_cau:
                      append_run_log(run_log, f"DUNG_EM thoat sau thang={thang}")
                      break

                # === Cho worker parse het queue ===
                # #print(f"\nCho worker parse het XML trong queue ({xml_queue.qsize()} con lai)...")
                # xml_queue.join()
                #print(f"\nCho worker parse het XML trong queue ({xml_queue.qsize()} con lai)...")
                append_run_log(run_log, f"QUEUE_JOIN_START remaining={xml_queue.qsize()}")
                heartbeat("Đang cho Worker parse hết XML", force=True, queue_size=xml_queue.qsize())
                xml_queue.join()
                append_run_log(run_log, "QUEUE_JOIN_DONE")

                stop_signal.set()
                xml_queue.put(None)
                worker.join(timeout=10)

                #print(f"Parse: {ket_qua_convert['ok']} OK, {ket_qua_convert['err']} loi")
                # === Bo sung cac dong trong Excel tong hop KHONG CO XML (chi ap dung cho file VAO_KM) ===
                append_run_log(run_log, f"NO_XML_APPEND_START excel_files={len(excel_summary_files)}")
                heartbeat("Đang bổ sung các dòng không có XML từ Excel tổng hợp", force=True)

                added_master_no_xml, added_line_no_xml = append_non_xml_rows_from_excel_files(
                    excel_file_infos=excel_summary_files,
                    all_master_rows=all_master_rows,
                    all_line_rows=all_line_rows,
                    lock_rows=lock_rows,
                    ma_donvi=MA_DONVI,
                    mst_tra_cuu=mst_value,
                    nam=nam,
                    run_log=run_log
                )
                # added_master_no_xml, added_line_no_xml = append_non_xml_rows_from_excel_files(
                #     excel_file_infos=excel_summary_files,
                #     all_master_rows=all_master_rows,
                #     all_line_rows=all_line_rows,
                #     lock_rows=lock_rows,
                #     run_log=run_log
                # )

                append_run_log(
                    run_log,
                    f"NO_XML_APPEND_DONE master={added_master_no_xml} line={added_line_no_xml}"
                )

                #print(f"Bo sung tu Excel khong XML: master={added_master_no_xml}, line={added_line_no_xml}")

                # === Cap nhat TTHAI_HD tu TAT CA file Excel tong hop (ca VAO va RA) ===
                append_run_log(run_log, f"TTHAI_HD_UPDATE_START excel_files={len(excel_summary_files)}")
                heartbeat("Đang cập nhật trạng thái hóa đơn từ Excel tổng hợp", force=True)

                updated_tthai = update_master_tthai_hd_from_excel_files(
                    excel_file_infos=excel_summary_files,
                    all_master_rows=all_master_rows,
                    lock_rows=lock_rows,
                    run_log=run_log
                )

                append_run_log(run_log, f"TTHAI_HD_UPDATE_DONE updated={updated_tthai}")
                #print(f"Cap nhat TTHAI_HD tu Excel tong hop: {updated_tthai} hoa don")                

                # === Ghi Excel + DBF theo huong ===
                excel_tong = None
                excel_tong_list = []
                stage_dir = paths["stage_dir"]
                os.makedirs(paths["output_dir"], exist_ok=True)

                # Dừng giữa chừng thì KHÔNG ghi Excel tổng. File này luôn được ghi ĐÈ
                # toàn bộ từ dữ liệu của lần chạy hiện tại, nên ghi lúc dở dang sẽ cho
                # ra một Excel tổng chỉ có vài tháng/vài loại — bước nạp đọc phải nó sẽ
                # tưởng đó là tất cả. Bỏ qua thì bản cũ còn nguyên vẹn, và lần chạy sau
                # dựng lại ĐỦ: XML đã tải vẫn được đẩy qua hàng đợi nhờ nhánh dung_lai.
                if dung_theo_yeu_cau:
                    append_run_log(run_log,
                        "BO_QUA_EXCEL_TONG da dung giua chung - giu nguyen ban cu")
                    status.write({**state, "state": "DA_DUNG",
                                  "message": "Đã dừng theo yêu cầu — đã chốt đường dẫn XML, "
                                             "chưa ghi Excel tổng"})
                elif all_master_rows or all_line_rows:
                    if loai_xuat == "all":
                        # Tach theo huong, ghi 2 file rieng
                        for h in ["RA", "VAO"]:
                            masters_h = [r for r in all_master_rows if r.get("HUONG") == h]
                            lines_h = [r for r in all_line_rows if r.get("HUONG") == h]
                            if not masters_h and not lines_h:
                                continue
                            excel_path = os.path.join(paths["output_dir"], f"HOA_DON_{h}_{MA_DONVI}.xlsx")
                            #print(f"Dang ghi {len(masters_h)} hoa don + {len(lines_h)} dong hang hoa vao: {excel_path}")
                            append_run_log(run_log, f"WRITE_EXCEL_START path={excel_path} master={len(masters_h)} lines={len(lines_h)}")
                            tong_m, tong_l = ghi_excel_tong(
                                excel_path, masters_h, lines_h,
                                f'hoa_don_{h.lower()}', f'hoa_don_{h.lower()}_line', run_log)
                            #print(f"Da ghi xong: {excel_path}")

                            events.log("EXCEL_TONG_WRITTEN", path=excel_path, master=tong_m, lines=tong_l, huong=h)
                            append_run_log(run_log, f"WRITE_EXCEL_DONE path={excel_path}")
                            append_run_log(run_log, f"TO_DBF_STAGE_START excel={excel_path} stage_dir={stage_dir}")
                            # excel_to_exact_dbf(excel_path, stage_dir)
                            append_run_log(run_log, f"TO_DBF_STAGE_DONE excel={excel_path}")
                            excel_tong_list.append(excel_path)
                                       
                        excel_tong = excel_tong_list
                    else:
                        # Chi 1 huong (ra hoac vao)
                        append_run_log(run_log, f"WRITE_EXCEL_START path={excel_tong} master={len(all_master_rows)} lines={len(all_line_rows)}")     
                        huong_label = loai_xuat.upper()
                        excel_tong = os.path.join(paths["output_dir"], f"HOA_DON_{huong_label}_{MA_DONVI}.xlsx")
                        append_run_log(run_log, f"WRITE_EXCEL_START path={excel_tong} master={len(all_master_rows)} lines={len(all_line_rows)}")
                        #print(f"Dang ghi {len(all_master_rows)} hoa don + {len(all_line_rows)} dong hang hoa vao: {excel_tong}")                        

                        # huong_label = loai_xuat.upper()
                        # excel_tong = os.path.join(paths["output_dir"], f"HOA_DON_{huong_label}_{MA_DONVI}.xlsx")
                        # #print(f"Dang ghi {len(all_master_rows)} hoa don + {len(all_line_rows)} dong hang hoa vao: {excel_tong}")

                        append_run_log(run_log, f"WRITE_EXCEL_DONE path={excel_tong}")
                        ghi_excel_tong(
                            excel_tong, all_master_rows, all_line_rows,
                            f'hoa_don_{huong_label.lower()}',
                            f'hoa_don_{huong_label.lower()}_line', run_log)

                        #print(f"Da ghi xong: {excel_tong}")
                        events.log("EXCEL_TONG_WRITTEN", path=excel_tong, master=len(all_master_rows), lines=len(all_line_rows))
                        append_run_log(run_log, f"TO_DBF_STAGE_START excel={excel_tong} stage_dir={stage_dir}")
                        # excel_to_exact_dbf(excel_tong, stage_dir)
                        append_run_log(run_log, f"TO_DBF_STAGE_DONE excel={excel_tong}")

                status.write({
                    **state,
                    "state": "PARSING_DONE",
                    "parse_ok": ket_qua_convert["ok"],
                    "parse_err": ket_qua_convert["err"],
                })
                status.write({
                    **state,
                    "state": "DONE_PARSE",
                    "message": "Đã tạo DBF stage",
                    "dbf_stage_path": stage_dir,
                    "ready_for_vfp_import": True
                })

                if ket_qua_convert["loi"]:
                    log_convert = os.path.join(paths["job_dir"], "LOI_PARSE.txt")
                    with open(log_convert, "w", encoding="utf-8") as f:
                        for ten, err in ket_qua_convert["loi"]:
                            f.write(f"{ten} - {err}\n")
                    events.log("PARSE_ERRORS", count=len(ket_qua_convert["loi"]), path=log_convert)

                status.write({
                    **state,
                    "state": "DONE_PARSE",
                    # Con so CHOT. Truoc day dong nay khong mang total/downloaded nen web
                    # giu nguyen anh chup giua chung ("10/29") du job da xong tu lau.
                    "total": state.get("tong_hd", 0),
                    "downloaded": state.get("tong_hd", 0),
                    "ok": state.get("tai_ok", 0),
                    "err": state.get("khong_co_goc", 0) + state.get("loi_that", 0),
                    "message": (f"Hoàn tất — {state.get('tai_ok', 0)}/{state.get('tong_hd', 0)} hóa đơn"
                                + (f", {state.get('khong_co_goc', 0)} không có hồ sơ gốc"
                                   if state.get("khong_co_goc", 0) else "")
                                + (f", {state.get('loi_that', 0)} lỗi cần xem lại"
                                   if state.get("loi_that", 0) else "")),
                    "alive": False,
                    "pid": os.getpid(),
                    "total_files": len(result),
                    "parse_ok": ket_qua_convert["ok"],
                    "parse_err": ket_qua_convert["err"],
                    "excel_tong": excel_tong,
                    "finished_at": datetime.datetime.now().isoformat(timespec="seconds")
                })
                
                events.log("JOB_DONE", total_files=len(result), parse_ok=ket_qua_convert["ok"], parse_err=ket_qua_convert["err"])
                append_run_log(run_log, f"JOB_DONE total_files={len(result)} parse_ok={ket_qua_convert['ok']} parse_err={ket_qua_convert['err']}")
                return result, excel_tong,{
                                                "status": "DONE_PARSE",
                                                "job_id": paths["job_id"],
                                                "stage_dir": paths["stage_dir"]
                                            }              
                  
                # status.write({**state, "state": "DONE_PARSE", "message": "Hoàn tất", "total_files": len(result), "parse_ok": ket_qua_convert["ok"], "parse_err": ket_qua_convert["err"], "excel_tong": excel_tong, "finished_at": datetime.datetime.now().isoformat(timespec="seconds")})
                # events.log("JOB_DONE", total_files=len(result), parse_ok=ket_qua_convert["ok"], parse_err=ket_qua_convert["err"])
                # return result, excel_tong,{
                #                                 "status": "DONE_PARSE",
                #                                 "job_id": paths["job_id"],
                #                                 "stage_dir": paths["stage_dir"]
                #                             }
        except Exception as e:
            import traceback
            tb = traceback.format_exc()

            #print(f"Loi: {e}\n{tb}")
            append_run_log(run_log, f"JOB_ERROR {short_exc(e)}")
            append_run_log(run_log, tb)

            try:
                driver.quit()
            except:
                pass

            try:
                shutil.rmtree(temp_profile_dir, ignore_errors=True)
            except:
                pass

            try:
                stop_signal.set()
                xml_queue.put(None)
            except:
                pass

            logging.error(f"Loi xay ra: {str(e)}", exc_info=True)

            try:
                status.write({
                    **state,
                    "state": "ERROR",
                    "message": str(e),
                    "alive": False,
                    "pid": os.getpid(),
                    "last_error": short_exc(e),
                    "finished_at": datetime.datetime.now().isoformat(timespec="seconds")
                })
                events.log("JOB_ERROR", error=str(e))
            except Exception:
                pass

            return f"error: {str(e)}"
        finally:
            if temp_profile_dir:
                shutil.rmtree(temp_profile_dir, ignore_errors=True)

        # except Exception as e:
        #     #print(f"Loi: {e}")
        #     try:
        #         driver.quit()
        #     except:
        #         pass
        #     shutil.rmtree(temp_profile_dir, ignore_errors=True)
        #     stop_signal.set()
        #     xml_queue.put(None)
        #     logging.error(f"Loi xay ra: {str(e)}", exc_info=True)
        #     try:
        #         status.write({**state, "state": "ERROR", "message": str(e), "finished_at": datetime.datetime.now().isoformat(timespec="seconds")})
        #         events.log("JOB_ERROR", error=str(e))
        #     except Exception:
        #         pass
        #     return f"error: {str(e)}"


# =========================
# CLI
# =========================
CLI_ERROR_LOG = os.path.join(_LOG_DIR, "cli_error.log")   # vá #7: không ghim ổ C

def _log_cli_error(msg):
    try:
        os.makedirs(os.path.dirname(CLI_ERROR_LOG), exist_ok=True)
        with open(CLI_ERROR_LOG, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.datetime.now().isoformat(timespec='seconds')}] {msg}\n")
            f.write(f"  argv: {' '.join(sys.argv)}\n")
    except Exception:
        pass

class LoggingArgumentParser(argparse.ArgumentParser):
    def error(self, message):
        _log_cli_error(f"argparse: {message}")
        super().error(message)

def parse_args(argv):
    p = LoggingArgumentParser(description="Xuat hoa don dien tu + convert XML->Excel song song.")
    p.add_argument("--run", action="store_true")
    p.add_argument("--mst", required=False)
    p.add_argument("--password", required=False)
    p.add_argument("--thang_bd", type=int, required=False)
    p.add_argument("--thang_kt", type=int, required=False)
    p.add_argument("--nam", type=int, required=False)
    p.add_argument("--save_dir", required=False)
    p.add_argument("--ma_donvi", required=False)
    p.add_argument("--status", required=False, help="Duong dan file JSON de ghi trang thai va tien do (co the dung de theo doi qua API tu xa)")
    p.add_argument("--events", required=False, help="Duong dan file JSONL de ghi log su kien (co the dung de theo doi qua API tu xa)")
    p.add_argument("--stagedir", required=False, help="Duong dan folder de luu DBF tam truoc khi import vao VFP")
    p.add_argument("--loai", default="all", choices=["all", "ra", "vao"])
    p.add_argument("--job_id", required=False, help="Ma duy nhat cho job, dung de tracking va luu tru ket qua theo job_id")
    p.add_argument("--to_dbf", action="store_true", help="Tu dong chuyen Excel tong sang DBF sau khi tai xong.")
    p.add_argument("--xml_map", required=False, help="Duong dan file XML_MAP.xlsx")
    p.add_argument("--tu_ngay", required=False, help="Tu ngay")
    p.add_argument("--den_ngay", required=False, help="Den ngay")
    # MAC DINH TAT. Khong truyen thi script chay y het truoc day, tung dong mot.
    # Day la dieu kien de thu che do moi ma khong danh cuoc luot chay that.
    p.add_argument("--tang_dan", action="store_true",
                   help="Bo tai lai nhung hoa don da co XML (hop nhat voi Excel lan truoc)")
    args = p.parse_args(argv)
    # Va #1: uu tien bien moi truong HDDT_PASSWORD (tham so dong lenh lo trong Task Manager)
    if not args.password:
        args.password = os.environ.get("HDDT_PASSWORD", "")
    return args
    # return p.parse_args(argv)

def run_cli(args):
    # Co gang tao run_log som nhat co the
    base_for_log = args.save_dir if args.save_dir else r"C:\test"

    ma_donvi_for_log = args.ma_donvi if args.ma_donvi else "UNKNOWN_DV"
    job_id_for_log = args.job_id if args.job_id else f"NOJOB_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}"
    # Dung CHUNG make_job_paths voi xuat_hoa_don, khong tu ghep duong dan o day.
    # Ban cu ghep tay <save_dir>\<MA>\<job_id>\run.log — THIEU tang NAM<nam>, nen moi
    # luot chay de lai mot thu muc chi chua dung mot file run.log nam canh NAM2026,
    # trong khi run.log that nam trong NAM2026\... Hai file cung ten, hai cho khac nhau.
    try:
        run_log = make_job_paths(base_for_log, job_id_for_log, ma_donvi_for_log)["run_log"]
    except Exception:
        # Ghi log som la de bat loi khoi dong — no khong duoc phep tu lam chet chuong trinh
        run_log = os.path.join(base_for_log, ma_donvi_for_log, job_id_for_log, "run.log")

    append_run_log(run_log, f"CLI_START pid={os.getpid()}")
    append_run_log(run_log, f"ARGV={' '.join(sys.argv)}")

    required = {
        "mst": args.mst, "password": args.password,
        "thang_bd": args.thang_bd, "thang_kt": args.thang_kt, "nam": args.nam,
        "save_dir": args.save_dir, "ma_donvi": args.ma_donvi, "job_id": args.job_id,
        "status": args.status, "events": args.events, "stagedir": args.stagedir
    }
    missing = [k for k, v in required.items() if not v]
    if missing:
        msg = f"Thieu tham so: {', '.join(missing)}"
        append_run_log(run_log, msg)
        #print(msg, file=sys.stderr)
        _log_cli_error(msg)
        sys.exit(2)

    tc = tra_cuu_hdt()

    try:
        append_run_log(run_log, "GO_INTO_XUAT_HOA_DON")
        res = tc.xuat_hoa_don(
            mst_value=args.mst,
            password_value=args.password,
            thang_bd=args.thang_bd,
            thang_kt=args.thang_kt,
            nam=args.nam,
            save_dir=args.save_dir,
            MA_DONVI=args.ma_donvi,
            loai_xuat=args.loai,
            job_id=args.job_id,
            status=args.status,
            events=args.events,
            stagedir=args.stagedir,
            xml_map_path=args.xml_map,
            tu_ngay=args.tu_ngay,
            den_ngay=args.den_ngay,
            tang_dan=args.tang_dan
        )

        append_run_log(run_log, f"XUAT_HOA_DON_RETURN {repr(res)[:1000]}")
        #print(f"Ket qua: {res}")

        if args.to_dbf and isinstance(res, tuple) and len(res) >= 2:
            excel_list = res[1]
            if not isinstance(excel_list, list):
                excel_list = [excel_list]
            for excel_path in excel_list:
                if excel_path and os.path.exists(excel_path):
                    out_dir = os.path.dirname(excel_path)
                    append_run_log(run_log, f"TO_DBF_START excel={excel_path}")
                    #print(f"Dang chuyen DBF: {excel_path} -> {out_dir}")
                    try:
                        # excel_to_exact_dbf(excel_path, out_dir)
                        append_run_log(run_log, f"TO_DBF_DONE excel={excel_path}")
                    except Exception as e:
                        append_run_log(run_log, f"TO_DBF_ERROR {short_exc(e)}")
                        #print(f"Loi chuyen DBF: {e}")

        append_run_log(run_log, "CLI_DONE_OK")

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        append_run_log(run_log, f"CLI_ERROR {short_exc(e)}")
        append_run_log(run_log, tb)
        raise




if __name__ == "__main__":
    try:
        if "--run" in sys.argv:

            args = parse_args(sys.argv[1:])
            run_cli(args)
            #lệnh chạy 
            # PS C:\Project\Python\PYTHON> python TRA_CUU_HDDT_1_3.py --run --mst 0107130530 --password Meva852015@c --thang_bd 3 --thang_kt 3 --nam 2023 --save_dir C:\test --ma_donvi USA_MEVA --loai vao --job_id 1234 --status C:\test\jobs\1234\status.json --events C:\test\jobs\1234\events.jsonl
        elif any(a in sys.argv for a in ["--register", "--unregister", "/regserver", "/unregserver"]):
            # Chi dang ky COM khi co arg tuong ung (tranh chay 2 lan do elevation)
            win32com.server.register.UseCommandLine(tra_cuu_hdt)
        else:
            MST = "0500445938"
            Password = "supT1aA@"
            Thang_bd = 1
            Thang_Kt = 1
            Nam = 2025
            Save_dir = r"C:\test"
            xml_map_path = r"\\Server-test\data_hddt\xml_map.xlsx"
            MA_DONVI = "HUY_THANH"
            Loai_xuat = "all"
            job_id = f"T{Thang_bd}_{Nam}_{MA_DONVI}"
            paths = make_job_paths(Save_dir, job_id,MA_DONVI)
            status = paths["status"]
            events = paths["events"]
            stagedir = paths["stage_dir"]

            result = tra_cuu_hdt.xuat_hoa_don(
                tra_cuu_hdt,
                MST, Password,None,None, Thang_bd, Thang_Kt, Nam, Save_dir, MA_DONVI, job_id,
                status=status,
                events=events,
                stagedir=stagedir,
                loai_xuat=Loai_xuat,
                xml_map_path=xml_map_path
            )            
            # result = tra_cuu_hdt.xuat_hoa_don(
            #     tra_cuu_hdt,
            #     MST, Password, Thang_bd, Thang_Kt, Nam, Save_dir, MA_DONVI, job_id,status=status, events=events,stagedir=stagedir,
            #     loai_xuat=Loai_xuat
            # )
            # excel_to_exact_dbf(result[1], Save_dir)
            #cấu trúc thư mục lưu kết quả:
            # Save_dir/{job_id}/raw/RA|VAO/HD_RA_Tx|HD_VAO_Tx.xlsx
            # Save_dir/{job_id}/stage/HOA_DON.dbf|HOA_DON_LINE.dbf
            # Save_dir/{job_id}/status.json
            # Save_dir/{job_id}/events.jsonl
            #print(f"Ket qua: {result}")
            # Tiến độ sẽ theo dõi qua {save_dir}\jobs\{job_id}\status.json và events.jsonl.
    except SystemExit:
        raise
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        #print(f"Loi xay ra: {e}", file=sys.stderr)
        _log_cli_error(f"exception: {e}\n{tb}") 
    finally:
        sys.exit(1)

        #python TRA_CUU_HDDT_1_3.py --run --mst "0107130530" --password "Meva852015@c" --thang_bd "4" --thang_kt "4" --nam "2026" --save_dir "\\SERVER-TEST\D\DATA_HDDT\USA_MEVA\jobs\HDDT_USA_MEVA_20260414150635\" --ma_donvi "USA_MEVA" --status "\\SERVER-TEST\D\DATA_HDDT\USA_MEVA\jobs\HDDT_USA_MEVA_20260414150635\status.json" --loai "vao"  --events "\\SERVER-TEST\D\DATA_HDDT\USA_MEVA\jobs\HDDT_USA_MEVA_20260414150635\events.jsonl"  --job_id HDDT_USA_MEVA_20260414150635 --stagedir "\\SERVER-TEST\D\DATA_HDDT\USA_MEVA\jobs\HDDT_USA_MEVA_20260414150635"
        #"C:\Project\Python\.venv\Scripts\python.exe" "C:\Project\Python\PYTHON\TRA_CUU_HDDT_1_3.py" --run --mst "5300670477" --password "eAzB1aA@" --thang_bd "4" --thang_kt "4" --nam "2026" --save_dir "\\SERVER-HYEN\DATA_HDDT\" --ma_donvi "VINH_HOAN" --status "\\SERVER-HYEN\DATA_HDDT\VINH_HOAN\T4_2026_VINH_HOAN\status.json" --loai "vao" --event "\\SERVER-HYEN\DATA_HDDT\VINH_HOAN\T4_2026_VINH_HOAN\events.jsonl" --job_id "T4_2026_VINH_HOAN" --stagedir "\\SERVER-HYEN\DATA_HDDT\VINH_HOAN\T4_2026_VINH_HOAN\stage\"